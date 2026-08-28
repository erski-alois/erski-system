from flask import Flask, request, jsonify, send_from_directory, Response, send_file
import functools
import os
import json

from db import get_conn, init_db, rows_to_dicts, NOW_SQL
import db as _db
import auth
import authtoken
import booking
import pricing
import payroll
import config

FRONTEND_DIR = os.path.join(os.path.dirname(__file__), "..", "frontend")
app = Flask(__name__)


# ------------------------------------------------------------------
# 排程:團課開課前24小時內未滿成班人數自動取消(對照規則書§三)
# ------------------------------------------------------------------
# 這支邏輯原本只有一個要「有人/有排程呼叫」才會執行的手動API
# (/api/admin/group-classes/check-auto-cancel),不是真的自動背景執行。
# 這裡加上APScheduler,讓它變成真的每隔一段時間自動跑一次。
#
# 正式環境用gunicorn多個worker process(見render.yaml的-w 2),如果每個
# worker都各自啟動一份排程,同一個檢查會被重複執行好幾次(邏輯本身雖然是
# 幂等的、重複執行不會造成資料錯誤，但會浪費資源、也可能讓通知重複寫入)。
# 用PostgreSQL的advisory lock確保同一時間只有一個worker真的執行這個工作;
# 本機SQLite開發模式下通常只有單一process,不需要這個保護就直接執行。
_GROUP_CLASS_AUTO_CANCEL_LOCK_KEY = 872634501  # 任意固定整數,只要在這個系統裡不跟其他advisory lock用途撞key即可


def _run_group_class_auto_cancel_job():
    if _db.USE_POSTGRES:
        lock_conn = _db.psycopg2.connect(_db.DATABASE_URL)
        try:
            lock_conn.autocommit = True
            cur = lock_conn.cursor()
            cur.execute("SELECT pg_try_advisory_lock(%s)", (_GROUP_CLASS_AUTO_CANCEL_LOCK_KEY,))
            got_lock = cur.fetchone()[0]
            if not got_lock:
                return  # 別的worker正在跑或剛跑完,這個worker這次不用做事
            try:
                cancelled = booking.check_and_auto_cancel_group_classes()
                if cancelled:
                    app.logger.info(f"[團課自動取消排程] 已取消場次: {cancelled}")
            finally:
                cur.execute("SELECT pg_advisory_unlock(%s)", (_GROUP_CLASS_AUTO_CANCEL_LOCK_KEY,))
        finally:
            lock_conn.close()
    else:
        cancelled = booking.check_and_auto_cancel_group_classes()
        if cancelled:
            app.logger.info(f"[團課自動取消排程] 已取消場次: {cancelled}")


def _start_scheduler_once():
    """避免Flask debug reloader(會啟動兩個process)或重複import造成排程被啟動兩次。"""
    if os.environ.get("WERKZEUG_RUN_MAIN") == "false":
        return  # debug reloader的第一個(父)process,不用啟動
    if getattr(app, "_erski_scheduler_started", False):
        return
    try:
        from apscheduler.schedulers.background import BackgroundScheduler
    except ImportError:
        app.logger.warning("未安裝APScheduler,團課自動取消排程不會執行,請執行 pip install -r requirements.txt")
        return
    scheduler = BackgroundScheduler(timezone="Asia/Taipei")
    # 每15分鐘檢查一次即可:規則是「開課前24小時內未滿成班人數」,不需要到分鐘級即時性
    scheduler.add_job(_run_group_class_auto_cancel_job, "interval", minutes=15, id="group_class_auto_cancel")
    scheduler.start()
    app._erski_scheduler_started = True
    app.logger.info("團課自動取消排程已啟動(每15分鐘檢查一次)")


if os.environ.get("ERSKI_DISABLE_SCHEDULER") != "1":
    _start_scheduler_once()


@app.route("/")
def serve_frontend():
    # index.html 內含所有前台邏輯(訂課/價格計算等),改動頻繁,絕對不能被瀏覽器快取,
    # 否則重新部署後使用者還是會看到舊版邏輯,以為改動沒生效。
    resp = send_from_directory(FRONTEND_DIR, "index.html")
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp


@app.route("/coach")
def serve_frontend_coach_entry():
    """教練專屬頁面的獨立網址(可以直接分享/加入書籤給教練,不用先經過會員首頁)。
    還是同一份index.html(單頁應用程式),前端JS會偵測網址是/coach、自動直接開啟教練登入畫面。"""
    return serve_frontend()


@app.route("/manifest.json")
def serve_manifest():
    return send_from_directory(FRONTEND_DIR, "manifest.json", mimetype="application/manifest+json")


@app.route("/sw.js")
def serve_service_worker():
    # Service Worker 必須從網站根目錄提供,且不能被瀏覽器快取太久,否則更新不會立即生效
    resp = send_from_directory(FRONTEND_DIR, "sw.js", mimetype="application/javascript")
    resp.headers["Cache-Control"] = "no-cache"
    return resp


@app.route("/icon-192.png")
def serve_icon_192():
    return send_from_directory(FRONTEND_DIR, "icon-192.png", mimetype="image/png")


@app.route("/icon-512.png")
def serve_icon_512():
    return send_from_directory(FRONTEND_DIR, "icon-512.png", mimetype="image/png")


@app.after_request
def add_cors_headers(response):
    response.headers["Access-Control-Allow-Origin"] = "*"
    # 2026-08:X-Staff-Id(未簽章、可偽造的純工號)全面改成X-Staff-Token/X-Member-Token
    # (見authtoken.py),詳見README_部署交接指南.md第二節「會員身分沒有真的驗證」。
    response.headers["Access-Control-Allow-Headers"] = "Content-Type, X-Staff-Token, X-Member-Token"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS"
    return response


@app.route("/api/<path:_any>", methods=["OPTIONS"])
def cors_preflight(_any):
    return "", 204


ROLE_RANK = {"coach": 1, "cs": 2, "manager": 3, "boss": 4}


# ------------------------------------------------------------------
# 身分驗證(2026-08新增):員工/會員登入後,前端要帶著簽章token發後續請求,
# 後端一律驗證token,不再相信前端直接聲稱的work_id/member_id。
# 員工token放在 X-Staff-Token header,會員token放在 X-Member-Token header
# (兩者互不影響,同一個瀏覽器分頁可以同時登入會員+員工,兩個token都會被送出)。
# ------------------------------------------------------------------
def _current_staff():
    """驗證X-Staff-Token,回傳員工資料dict;沒帶token/token無效過期/帳號停用一律回傳None。
    不會中斷請求——給「員工或本人都可以」的路由自行判斷用;需要強制要求員工身分的路由請用require_role。"""
    work_id = authtoken.verify_staff_token(request.headers.get("X-Staff-Token"))
    if not work_id:
        return None
    conn = get_conn()
    staff = conn.execute("SELECT * FROM staff WHERE work_id=?", (work_id,)).fetchone()
    conn.close()
    if not staff or not staff["is_active"]:
        return None
    return dict(staff)


def _current_member_id():
    """驗證X-Member-Token,回傳token內的member_id(int);沒帶token/token無效過期回傳None。"""
    return authtoken.verify_member_token(request.headers.get("X-Member-Token"))


def require_role(min_role):
    def decorator(f):
        @functools.wraps(f)
        def wrapper(*args, **kwargs):
            staff = _current_staff()
            if not staff:
                return jsonify({"error": "未登入或登入已過期,請重新登入"}), 401
            if ROLE_RANK.get(staff["role"], 0) < ROLE_RANK[min_role]:
                return jsonify({"error": "權限不足"}), 403
            request.current_staff = staff
            return f(*args, **kwargs)
        return wrapper
    return decorator


def require_member_or_staff(min_staff_role="cs"):
    """給/api/members/<member_id>/...這類路由用:必須是「本人的會員token」,
    或是「min_staff_role以上的員工token」(員工可以查看/操作任何會員的資料,對照現有
    客服後台「查看會員詳情」功能),不再直接信任路徑上的member_id參數。"""
    def decorator(f):
        @functools.wraps(f)
        def wrapper(member_id, *args, **kwargs):
            staff = _current_staff()
            if staff and ROLE_RANK.get(staff["role"], 0) >= ROLE_RANK[min_staff_role]:
                request.current_staff = staff
                return f(member_id, *args, **kwargs)
            auth_member_id = _current_member_id()
            if auth_member_id is None:
                return jsonify({"error": "未登入或登入已過期,請重新登入"}), 401
            if auth_member_id != member_id:
                return jsonify({"error": "無權限操作其他會員的資料"}), 403
            return f(member_id, *args, **kwargs)
        return wrapper
    return decorator


def _require_member_id_from_token():
    """給訂課/付款這類路由用:一律用X-Member-Token裡的member_id,不再信任前端傳來的
    member_id欄位(避免有人把body裡的member_id改成別人的id去下單/查詢)。
    回傳(member_id, None)表示驗證成功;驗證失敗回傳(None, (response, status_code)),
    呼叫端直接 `return err` 即可。"""
    member_id = _current_member_id()
    if member_id is None:
        return None, (jsonify({"error": "未登入或登入已過期,請重新登入"}), 401)
    return member_id, None


# ------------------------------------------------------------------
# 會員登入 / 建立
# ------------------------------------------------------------------
@app.route("/api/auth/oauth-login", methods=["POST"])
def oauth_login():
    data = request.json
    result = auth.mock_oauth_login(data["provider"], data["mock_external_id"])
    # 找到既有會員=視同這次OAuth登入完成身分驗證(正式串接LINE/Google OAuth後,
    # 「第三方回傳的身分已通過驗證」這件事本來就是由OAuth provider保證,
    # 目前是mock_external_id模擬這個結果),核發一組會員token給前端後續請求使用。
    if not result.get("is_new") and result.get("member"):
        result["token"] = authtoken.issue_member_token(result["member"]["id"])
    return jsonify(result)


@app.route("/api/auth/create-member", methods=["POST"])
def create_member():
    member = auth.create_member(request.json)
    conn = get_conn()
    conn.execute(
        """INSERT INTO notifications (member_id, channel, notify_type, content, status)
           VALUES (?, 'email', 'registration_success', ?, 'simulated')""",
        (member["id"], f"歡迎加入 ERSKI,{member['name']} 您的會員註冊已完成!"),
    )
    conn.commit()
    conn.close()
    member.pop("password_hash", None)
    member["token"] = authtoken.issue_member_token(member["id"])
    return jsonify(member), 201


import random
import string


def _generate_partner_code():
    return "".join(random.choices(string.ascii_uppercase + string.digits, k=6))


@app.route("/api/admin/partners", methods=["GET"])
@require_role("manager")
def admin_list_partners():
    conn = get_conn()
    rows = conn.execute("SELECT * FROM partner_organizations ORDER BY created_at DESC").fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/admin/partners", methods=["POST"])
@require_role("manager")
def admin_create_partner():
    d = request.json
    conn = get_conn()
    code = d.get("code") or _generate_partner_code()
    try:
        cur = conn.execute(
            """INSERT INTO partner_organizations (name, contact_name, contact_phone, contact_email, code, rebate_rate)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (d["name"], d.get("contact_name"), d.get("contact_phone"), d.get("contact_email"),
             code, d.get("rebate_rate", 0)),
        )
        conn.commit()
        partner_id = cur.lastrowid
    except Exception:
        conn.close()
        return jsonify({"error": "此優惠碼已存在,請換一組"}), 400
    conn.close()
    return jsonify({"id": partner_id, "code": code}), 201


@app.route("/api/admin/partners/<int:partner_id>", methods=["PUT"])
@require_role("manager")
def admin_update_partner(partner_id):
    d = request.json
    conn = get_conn()
    conn.execute(
        "UPDATE partner_organizations SET name=?, contact_name=?, contact_phone=?, contact_email=?, rebate_rate=?, is_active=? WHERE id=?",
        (d["name"], d.get("contact_name"), d.get("contact_phone"), d.get("contact_email"),
         d.get("rebate_rate", 0), d.get("is_active", 1), partner_id),
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/partners/<int:partner_id>/report", methods=["GET"])
@require_role("manager")
def admin_partner_report(partner_id):
    """匯出該合作單位所推薦會員的訂單資料,供計算回饋/回扣參考。"""
    conn = get_conn()
    partner = conn.execute("SELECT * FROM partner_organizations WHERE id=?", (partner_id,)).fetchone()
    if not partner:
        conn.close()
        return jsonify({"error": "找不到此合作單位"}), 404
    members = conn.execute(
        "SELECT id, name, phone, email, created_at FROM members WHERE referral_partner_id=?", (partner_id,)
    ).fetchall()
    member_ids = [m["id"] for m in members]
    total_revenue = 0
    orders_detail = []
    if member_ids:
        placeholders = ",".join("?" * len(member_ids))
        # 2026-08確認:分潤只算「這位被推薦會員的第一筆已付款訂單」,不是這位會員
        # 之後所有消費都持續算分潤(常見於「介紹新客戶」型的一次性推薦獎勵,而非
        # 長期抽成),所以用相關子查詢限定只取每位會員最早的一筆paid訂單。
        orders = conn.execute(
            f"""SELECT o.*, m.name AS member_name FROM orders o JOIN members m ON o.member_id = m.id
                WHERE o.member_id IN ({placeholders}) AND o.status='paid'
                AND o.id = (
                    SELECT o2.id FROM orders o2
                    WHERE o2.member_id = o.member_id AND o2.status='paid'
                    ORDER BY o2.created_at ASC, o2.id ASC LIMIT 1
                )""",
            member_ids,
        ).fetchall()
        orders_detail = rows_to_dicts(orders)
        total_revenue = sum(o["amount"] for o in orders)
    conn.close()
    rebate_amount = round(total_revenue * (partner["rebate_rate"] or 0))
    return jsonify({
        "partner": dict(partner),
        "referred_members": rows_to_dicts(members),
        "referred_member_count": len(members),
        "total_paid_revenue": total_revenue,
        "rebate_rate": partner["rebate_rate"],
        "rebate_amount": rebate_amount,
        "rebate_basis": "每位推薦會員只計算第一筆已付款訂單(一次性推薦獎勵,非長期抽成);撥款需你自行手動處理,系統僅提供報表",
        "orders": orders_detail,
    })


@app.route("/api/members/<int:member_id>/referral-code", methods=["PUT"])
@require_member_or_staff()
def set_member_referral_code(member_id):
    """會員填寫優惠碼(僅能設定一次,設定過後不可再自行更改,避免濫用回饋歸屬)。"""
    d = request.json
    conn = get_conn()
    member = conn.execute("SELECT * FROM members WHERE id=?", (member_id,)).fetchone()
    if not member:
        conn.close()
        return jsonify({"error": "找不到此會員"}), 404
    if member["referral_code"]:
        conn.close()
        return jsonify({"error": "已經填寫過優惠碼,無法再次更改"}), 400
    partner = conn.execute(
        "SELECT * FROM partner_organizations WHERE code=? AND is_active=1", (d["code"],)
    ).fetchone()
    if not partner:
        conn.close()
        return jsonify({"error": "優惠碼不存在或已失效"}), 400
    conn.execute(
        "UPDATE members SET referral_code=?, referral_partner_id=? WHERE id=?",
        (d["code"], partner["id"], member_id),
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "partner_name": partner["name"]})


@app.route("/api/members/<int:member_id>/password", methods=["PUT"])
@require_member_or_staff()
def set_member_password(member_id):
    d = request.json
    try:
        auth.set_member_password(member_id, d["new_password"], d.get("current_password"))
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    conn = get_conn()
    m = conn.execute("SELECT name FROM members WHERE id=?", (member_id,)).fetchone()
    conn.execute(
        """INSERT INTO notifications (member_id, channel, notify_type, content, status)
           VALUES (?, 'email', 'password_changed', ?, 'simulated')""",
        (member_id, f"{m['name']} 您好,你的 ERSKI 帳號登入密碼已成功變更。若非本人操作請立即聯繫客服。"),
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/auth/staff-login", methods=["POST"])
def staff_login_route():
    data = request.json
    staff = auth.staff_login(data["work_id"], data["password"])
    if not staff:
        return jsonify({"error": "工號或密碼錯誤"}), 401
    staff["token"] = authtoken.issue_staff_token(staff["work_id"])
    return jsonify(staff)


@app.route("/api/members/<int:member_id>", methods=["GET"])
@require_member_or_staff()
def get_member(member_id):
    conn = get_conn()
    m = conn.execute("SELECT * FROM members WHERE id=?", (member_id,)).fetchone()
    passes = conn.execute(
        """SELECT cp.*, r.id AS pending_request_id, r.request_type AS pending_request_type,
                  r.requested_package_size AS pending_requested_package_size
           FROM charter_passes cp
           LEFT JOIN charter_pass_requests r ON r.charter_pass_id = cp.id AND r.status='pending'
           WHERE cp.member_id=?""",
        (member_id,),
    ).fetchall()
    plan = conn.execute(
        "SELECT * FROM member_plans WHERE member_id=? AND is_active=1", (member_id,)
    ).fetchone()
    conn.close()
    if not m:
        return jsonify({"error": "member not found"}), 404
    from datetime import date
    today = booking.today_tw().isoformat()
    quota = booking.get_quota_status(member_id, today)

    all_bookings = booking.get_all_bookings(member_id=member_id)
    my_bookings = [b for b in all_bookings if b["status"] != "cancelled"]
    # 保留原本 upcoming/incomplete 欄位供舊版相容(以付款狀態判斷,而非課程狀態)
    upcoming = [b for b in my_bookings if b["payment_status"] == "confirmed"]
    incomplete = [b for b in my_bookings if b["payment_status"] != "confirmed"]
    japan_payment_status = booking.get_member_japan_payment_status(member_id)

    m_dict = dict(m)
    has_password = bool(m_dict.pop("password_hash", None))

    return jsonify({
        "member": m_dict,
        "has_password": has_password,
        "member_code": compute_member_code(dict(m)),
        "profile_complete": auth.is_profile_complete(m),
        "charter_passes": rows_to_dicts(passes),
        "plan": dict(plan) if plan else None,
        "quota_status": quota,
        "my_bookings": my_bookings,
        "upcoming_bookings": upcoming,
        "incomplete_bookings": incomplete,
        "japan_payment_status": japan_payment_status,
    })


def compute_member_code(m):
    """會員編號 = 註冊年(4碼)+月(2碼)+滑行項目(ski=1/snowboard=2)+性別(男=1/女=2)+流水序號(4碼,用會員id補零)。
       性別或主要滑行項目尚未填寫時回傳 None(前端顯示尚未產生編號)。"""
    if not m.get("gender") or not m.get("primary_equipment") or not m.get("created_at"):
        return None
    year = m["created_at"][0:4]
    month = m["created_at"][5:7]
    equip_digit = "1" if m["primary_equipment"] == "ski" else "2"
    gender_digit = "1" if m["gender"] == "male" else "2"
    seq = str(m["id"]).zfill(4)
    return f"{year}{month}{equip_digit}{gender_digit}{seq}"


@app.route("/api/members/<int:member_id>/notifications", methods=["GET"])
@require_member_or_staff()
def member_notifications(member_id):
    conn = get_conn()
    rows = conn.execute(
        "SELECT * FROM notifications WHERE member_id=? ORDER BY created_at DESC LIMIT 50", (member_id,)
    ).fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/members/<int:member_id>/companions", methods=["GET"])
@require_member_or_staff()
def list_member_companions(member_id):
    conn = get_conn()
    rows = conn.execute(
        "SELECT * FROM member_companions WHERE member_id=? ORDER BY created_at DESC", (member_id,)
    ).fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/members/<int:member_id>/companions", methods=["POST"])
@require_member_or_staff()
def create_member_companion(member_id):
    d = request.json
    conn = get_conn()
    cur = conn.execute(
        """INSERT INTO member_companions (member_id, name, gender, age, height_cm, weight_kg, shoe_size, equipment_type)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
        (member_id, d.get("name"), d.get("gender"), d.get("age"), d.get("height_cm"),
         d.get("weight_kg"), d.get("shoe_size"), d.get("equipment_type")),
    )
    conn.commit()
    companion_id = cur.lastrowid
    conn.close()
    return jsonify({"id": companion_id}), 201


@app.route("/api/members/<int:member_id>/companions/<int:companion_id>", methods=["PUT"])
@require_member_or_staff()
def update_member_companion(member_id, companion_id):
    d = request.json
    conn = get_conn()
    conn.execute(
        """UPDATE member_companions SET name=?, gender=?, age=?, height_cm=?, weight_kg=?, shoe_size=?, equipment_type=?
           WHERE id=? AND member_id=?""",
        (d.get("name"), d.get("gender"), d.get("age"), d.get("height_cm"),
         d.get("weight_kg"), d.get("shoe_size"), d.get("equipment_type"), companion_id, member_id),
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/members/<int:member_id>/companions/<int:companion_id>", methods=["DELETE"])
@require_member_or_staff()
def delete_member_companion(member_id, companion_id):
    conn = get_conn()
    conn.execute("DELETE FROM member_companions WHERE id=? AND member_id=?", (companion_id, member_id))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/members/<int:member_id>/entitlement-ledger", methods=["GET"])
@require_member_or_staff()
def member_entitlement_ledger(member_id):
    conn = get_conn()
    rows = conn.execute(
        "SELECT * FROM entitlement_ledger WHERE member_id=? ORDER BY created_at DESC LIMIT 50", (member_id,)
    ).fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/members/<int:member_id>/profile", methods=["PUT"])
@require_member_or_staff()
def update_member_profile(member_id):
    d = request.json
    fields = [
        "name", "phone", "emergency_contact_name", "emergency_contact_phone",
        "birth_date", "gender", "id_number", "blood_type", "address",
        "line_id", "social_handle", "height_cm", "weight_kg",
        "snowboard_length", "snowboard_boot_size", "ski_length", "ski_boot_size",
        "machine_level", "snow_level", "primary_equipment",
    ]
    updates = {k: v for k, v in d.items() if k in fields}
    if not updates:
        return jsonify({"error": "no valid fields"}), 400
    # 2026-08:gender/primary_equipment欄位在資料庫有CHECK限制(gender只能是male/female,
    # primary_equipment只能是ski/snowboard),但前端下拉選單原本就允許選「未填寫/未選擇」
    # (對應空字串"")——這是正常、預期會出現的狀態(會員本來就可以先不填性別/主要滑行項目)。
    # 空字串不符合CHECK限制,原本會讓這裡的UPDATE直接丟出未攔截的sqlite3.IntegrityError,
    # 導致「儲存我的資料」整個500失敗,而且這個例外沒有被任何except接住,conn也沒有被
    # close(),已經實測會導致連線/寫入鎖沒釋放、拖累後續其他人的請求(database is locked)。
    # 修法:空字串一律當成「這個欄位還沒填」,存成NULL,不是直接把空字串寫進去。
    for k in ("gender", "primary_equipment"):
        if k in updates and updates[k] == "":
            updates[k] = None
    set_clause = ", ".join(f"{k}=?" for k in updates)
    conn = get_conn()
    try:
        conn.execute(f"UPDATE members SET {set_clause} WHERE id=?", (*updates.values(), member_id))
        conn.commit()
        row = conn.execute("SELECT * FROM members WHERE id=?", (member_id,)).fetchone()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    row_dict = dict(row)
    row_dict.pop("password_hash", None)  # 2026-08修正:原本這裡會把密碼雜湊值一起回傳給前端
    return jsonify(row_dict)


# ------------------------------------------------------------------
# 價目表(前台課表顯示用)
# ------------------------------------------------------------------
@app.route("/api/admin/pricing-config", methods=["GET"])
@require_role("manager")
def admin_list_pricing_config():
    return jsonify(pricing.list_all_configs())


@app.route("/api/admin/pricing-config/<config_key>", methods=["PUT"])
@require_role("manager")
def admin_update_pricing_config(config_key):
    d = request.json
    conn = get_conn()
    before = conn.execute("SELECT * FROM pricing_config WHERE config_key=?", (config_key,)).fetchone()
    if not before:
        conn.close()
        return jsonify({"error": "找不到此設定項目"}), 404
    try:
        json.loads(d["config_value"]) if isinstance(d["config_value"], str) else d["config_value"]
    except (ValueError, TypeError):
        conn.close()
        return jsonify({"error": "格式不是合法的JSON,請檢查輸入內容"}), 400
    conn.close()

    new_value = d["config_value"] if not isinstance(d["config_value"], str) else json.loads(d["config_value"])
    pricing.set_config(config_key, new_value, staff_id=request.current_staff["id"])

    conn2 = get_conn()
    conn2.execute(
        """INSERT INTO audit_log (staff_id, action, target_type, target_id, before_value, after_value)
           VALUES (?, 'update_pricing_config', 'pricing_config', ?, ?, ?)""",
        (request.current_staff["id"], before["id"], json.dumps({config_key: before["config_value"]}),
         json.dumps({config_key: new_value})),
    )
    conn2.commit()
    conn2.close()
    return jsonify({"ok": True})


@app.route("/api/pricing", methods=["GET"])
def get_pricing():
    return jsonify({
        "trial": pricing.get_config("trial_price"),
        "charter": pricing.get_config("charter_price"),
        "self_practice": pricing.get_config("self_practice_price"),
        "jump": pricing.get_config("jump_price"),
        "japan_full_day": pricing.get_config("japan_full_day_price"),
        "japan_half_day": pricing.get_config("japan_half_day_price"),
        "japan_designate_coach_fee": pricing.get_config("japan_coach_designate_fee"),
        "charter_designate_coach_fee": pricing.get_config("charter_coach_designate_fee"),
        "group_class_min": pricing.get_config("group_class_min"),
        "group_class_max": pricing.get_config("group_class_max"),
        "indoor_hours": {"start": pricing.get_config("indoor_start_hour"), "last_start": pricing.get_config("indoor_last_start_hour")},
        "plan_quota": pricing.get_config("plan_quota"),
        "plan_fee": pricing.get_config("plan_fee"),
        "booking_window_days": pricing.get_config("booking_window_days"),
        "japan_season_months": sorted(pricing.JAPAN_SEASON_MONTHS),
    })


# ------------------------------------------------------------------
# 室內滑雪:日曆可用性(某月/某日的機台狀況 + 跳台預約)
# ------------------------------------------------------------------
@app.route("/api/indoor/month/<year_month>", methods=["GET"])
def indoor_month_view(year_month):
    """year_month格式 YYYY-MM,回傳當月每天的機台/團課/跳台概況,供大日曆顯示狀態用。"""
    conn = get_conn()
    sessions = conn.execute(
        """SELECT booking_date, category, start_hour, duration_minutes, status FROM indoor_sessions
           WHERE booking_date LIKE ? AND status != 'cancelled'""",
        (f"{year_month}-%",),
    ).fetchall()
    jumps = conn.execute(
        "SELECT booking_date, start_time, duration_minutes FROM jump_bookings WHERE booking_date LIKE ? AND status != 'cancelled'",
        (f"{year_month}-%",),
    ).fetchall()
    by_day = {}
    for s in sessions:
        by_day.setdefault(s["booking_date"], {"indoor": 0, "jump": 0, "group_class": 0})
        if s["category"] == "group_class":
            by_day[s["booking_date"]]["group_class"] += 1
        else:
            by_day[s["booking_date"]]["indoor"] += 1
    for j in jumps:
        by_day.setdefault(j["booking_date"], {"indoor": 0, "jump": 0, "group_class": 0})
        by_day[j["booking_date"]]["jump"] += 1
    conn.close()
    return jsonify(by_day)


@app.route("/api/indoor/day/<date_str>", methods=["GET"])
def indoor_day_view(date_str):
    conn = get_conn()
    sessions = conn.execute(
        """SELECT * FROM indoor_sessions WHERE booking_date=?
           AND status != 'cancelled' ORDER BY start_hour""",
        (date_str,),
    ).fetchall()
    result = []
    for s in sessions:
        members = conn.execute(
            "SELECT * FROM indoor_session_members WHERE session_id=? AND status='enrolled'",
            (s["id"],),
        ).fetchall()
        result.append({**dict(s), "members": rows_to_dicts(members)})
    jumps = conn.execute(
        "SELECT * FROM jump_bookings WHERE booking_date=? AND status != 'cancelled'",
        (date_str,),
    ).fetchall()
    conn.close()
    return jsonify({"indoor_sessions": result, "jump_bookings": rows_to_dicts(jumps)})


# ------------------------------------------------------------------
# 體驗課
# ------------------------------------------------------------------
@app.route("/api/booking/trial", methods=["POST"])
def book_trial():
    member_id, err = _require_member_id_from_token()
    if err:
        return err
    d = request.json
    try:
        result = booking.book_trial(
            member_id=member_id, booking_date=d["booking_date"],
            start_hour=d["start_hour"], headcount=d["headcount"],
            equipment_type=d.get("equipment_type"), participants=d.get("participants"),
            coach_id=d.get("coach_id"),
        )
        return jsonify(result), 201
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


# ------------------------------------------------------------------
# 包機課(先購買堂數包,再用堂數包預約時段)
# ------------------------------------------------------------------
@app.route("/api/charter/purchase", methods=["POST"])
def purchase_charter():
    member_id, err = _require_member_id_from_token()
    if err:
        return err
    d = request.json
    try:
        result = booking.purchase_charter_pass(
            member_id=member_id, package_size=d["package_size"], headcount_type=d["headcount_type"],
            equipment_type=d.get("equipment_type"),
        )
        return jsonify(result), 201
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/booking/charter", methods=["POST"])
def book_charter():
    member_id, err = _require_member_id_from_token()
    if err:
        return err
    d = request.json
    try:
        result = booking.book_charter(
            member_id=member_id, booking_date=d["booking_date"], start_hour=d["start_hour"],
            charter_pass_id=d["charter_pass_id"],
            equipment_type=d.get("equipment_type"), participants=d.get("participants"),
            coach_id=d.get("coach_id"),
        )
        return jsonify(result), 201
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/indoor-coaches", methods=["GET"])
def list_indoor_coaches_public():
    """供客戶端選擇包機課「指定教練」時使用(不需要員工權限),
    列出後台「教練管理」裡駐在地被標記為室內滑雪分店的教練。"""
    conn = get_conn()
    rows = conn.execute(
        """SELECT DISTINCT s.id, s.name, s.display_code FROM coach_locations cl
           JOIN coach_location_options lo ON cl.location_option_id = lo.id
           JOIN staff s ON cl.coach_id = s.id
           WHERE lo.is_indoor_branch = 1 AND s.is_active = 1"""
    ).fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/members/<int:member_id>/charter-passes/<int:pass_id>/request", methods=["POST"])
@require_member_or_staff()
def member_request_charter_pass_change(member_id, pass_id):
    """會員對已購買的包機堂數包送出「取消」或「換堂數包大小」申請,
    退不退款/換多少堂,由客服後台審核決定(見admin_resolve_charter_pass_request)。"""
    d = request.json
    try:
        result = booking.request_charter_pass_change(
            member_id=member_id, charter_pass_id=pass_id,
            request_type=d["request_type"], requested_package_size=d.get("requested_package_size"),
            note=d.get("note"),
        )
        return jsonify(result), 201
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/admin/charter-pass-requests/pending", methods=["GET"])
@require_role("cs")
def admin_list_charter_pass_requests():
    return jsonify(booking.list_charter_pass_requests(status="pending"))


@app.route("/api/admin/charter-pass-requests/<int:request_id>/resolve", methods=["POST"])
@require_role("cs")
def admin_resolve_charter_pass_request(request_id):
    d = request.json
    try:
        result = booking.resolve_charter_pass_request(
            request_id=request_id, action=d["action"], staff_id=request.current_staff["id"],
            new_remaining=d.get("new_remaining"), staff_note=d.get("staff_note"),
        )
        return jsonify(result)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


# ------------------------------------------------------------------
# 自主練習
# ------------------------------------------------------------------
@app.route("/api/booking/self-practice", methods=["POST"])
def book_self_practice():
    member_id, err = _require_member_id_from_token()
    if err:
        return err
    d = request.json
    try:
        result = booking.book_self_practice(
            member_id=member_id, booking_date=d["booking_date"], start_hour=d["start_hour"],
            duration_minutes=d["duration_minutes"], headcount=d.get("headcount", 1),
            equipment_type=d.get("equipment_type"), participants=d.get("participants"),
            use_plan_quota=d.get("use_plan_quota", False),
        )
        return jsonify(result), 201
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


# ------------------------------------------------------------------
# 團課
# ------------------------------------------------------------------
@app.route("/api/booking/group-class", methods=["POST"])
def enroll_group_class():
    member_id, err = _require_member_id_from_token()
    if err:
        return err
    d = request.json
    try:
        result = booking.enroll_group_class(
            member_id=member_id, booking_date=d["booking_date"], start_hour=d["start_hour"],
            equipment_type=d.get("equipment_type"), participant=d.get("participant"),
        )
        return jsonify(result), 201
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


# ------------------------------------------------------------------
# 跳台體驗
# ------------------------------------------------------------------
@app.route("/api/booking/jump", methods=["POST"])
def book_jump():
    member_id, err = _require_member_id_from_token()
    if err:
        return err
    d = request.json
    try:
        result = booking.book_jump(
            member_id=member_id, booking_date=d["booking_date"],
            start_time=d["start_time"], duration_minutes=d["duration_minutes"],
            equipment_type=d.get("equipment_type"), participants=d.get("participants"),
        )
        return jsonify(result), 201
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


# ------------------------------------------------------------------
# 日本教練課
# ------------------------------------------------------------------
@app.route("/api/japan-regions", methods=["GET"])
def list_japan_regions():
    """公開API:回傳所有分區,每個分區附帶該分區的雪場清單(若需要選擇雪場)。"""
    conn = get_conn()
    regions = conn.execute(
        "SELECT * FROM japan_regions ORDER BY display_order"
    ).fetchall()
    result = []
    for r in regions:
        resorts = conn.execute(
            "SELECT * FROM ski_resorts WHERE region_id=? AND is_active=1", (r["id"],)
        ).fetchall()
        result.append({**dict(r), "resorts": rows_to_dicts(resorts)})
    conn.close()
    return jsonify(result)


@app.route("/api/resorts/<int:resort_id>/month/<year_month>", methods=["GET"])
def resort_month_view(resort_id, year_month):
    """year_month格式YYYY-MM,回傳該雪場當月每天「是否已有會員訂課」,供日本滑雪預約的
    日曆畫橘點用(跟室內滑雪/api/indoor/month同樣的概念):不代表已經額滿,只是提示當天
    已經有其他預約,實際能不能訂要看送出後系統依駐在教練數判斷的結果。"""
    conn = get_conn()
    rows = conn.execute(
        """SELECT booking_date, COUNT(*) AS cnt FROM japan_bookings
           WHERE resort_id=? AND booking_date LIKE ? AND status != 'cancelled'
           GROUP BY booking_date""",
        (resort_id, f"{year_month}-%"),
    ).fetchall()
    conn.close()
    return jsonify({r["booking_date"]: r["cnt"] for r in rows})


@app.route("/api/resorts", methods=["GET"])
def list_resorts():
    conn = get_conn()
    region_id = request.args.get("region_id", type=int)
    q = "SELECT * FROM ski_resorts WHERE is_active=1"
    params = []
    if region_id:
        q += " AND region_id=?"
        params.append(region_id)
    rows = conn.execute(q, params).fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/resorts/<int:resort_id>/coaches", methods=["GET"])
def list_resort_coaches_public(resort_id):
    """供客戶端選擇日本教練課指定教練時使用(不需要員工權限)。"""
    conn = get_conn()
    rows = conn.execute(
        """SELECT s.id, s.name, s.display_code FROM resort_coaches rc
           JOIN staff s ON rc.coach_id = s.id WHERE rc.resort_id=?""",
        (resort_id,),
    ).fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/admin/resorts/<int:resort_id>", methods=["DELETE"])
@require_role("manager")
def admin_delete_resort(resort_id):
    conn = get_conn()
    conn.execute("UPDATE ski_resorts SET is_active=0 WHERE id=?", (resort_id,))
    conn.execute(
        """INSERT INTO audit_log (staff_id, action, target_type, target_id, before_value, after_value)
           VALUES (?, 'delete_resort', 'ski_resort', ?, '{"is_active": 1}', '{"is_active": 0}')""",
        (request.current_staff["id"], resort_id),
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/coaches/export.xlsx", methods=["GET"])
@require_role("manager")
def admin_export_coaches_xlsx():
    """匯出全部教練完整資料(基本資料+教練介紹+能力證照+人事機密欄位)為Excel,僅限主管以上使用。"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    import tempfile

    conn = get_conn()
    coaches = conn.execute(
        """SELECT s.*, cp.rank, cp.resume, cp.experience, cp.self_intro, cp.contract_type,
                  cp.years_of_service, cp.contract_year, cp.hourly_rate, cp.base_salary,
                  cp.rate_group_class, cp.rate_trial, cp.rate_assistant, cp.japan_commission_rate
           FROM staff s LEFT JOIN coach_profiles cp ON cp.coach_id = s.id
           WHERE s.role='coach' ORDER BY s.name"""
    ).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "教練資料"
    headers = ["工號", "姓名", "電話", "分店", "在職狀態", "職稱", "資歷(年)", "經歷", "自我介紹",
               "合約類型", "年資", "合約年", "堂課時薪", "體驗時薪", "助教時薪", "基本薪資",
               "日本教練課提成比例", "教練能力", "證照", "駐在地"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E2761")

    for c in coaches:
        caps = conn.execute(
            """SELECT co.name FROM coach_capabilities cc
               JOIN coach_capability_options co ON cc.capability_option_id = co.id
               WHERE cc.coach_id=?""", (c["id"],),
        ).fetchall()
        certs = conn.execute(
            "SELECT cert_type, cert_name, cert_level FROM coach_certifications WHERE coach_id=?", (c["id"],)
        ).fetchall()
        locs = conn.execute(
            """SELECT lo.name FROM coach_locations cl
               JOIN coach_location_options lo ON cl.location_option_id = lo.id
               WHERE cl.coach_id=?""", (c["id"],),
        ).fetchall()
        ws.append([
            c["work_id"], c["name"], c["phone"], c["branch"], "在職" if c["is_active"] else "已停用",
            c["rank"], c["resume"], c["experience"], c["self_intro"],
            c["contract_type"], c["years_of_service"], c["contract_year"],
            c["rate_group_class"], c["rate_trial"], c["rate_assistant"], c["base_salary"],
            c["japan_commission_rate"],
            "、".join(r["name"] for r in caps),
            "、".join(f"{r['cert_name']}({r['cert_level']})" for r in certs),
            "、".join(r["name"] for r in locs),
        ])
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)
    conn.close()

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        wb.save(tmp.name)
        path = tmp.name
    return send_file(path, as_attachment=True, download_name="教練資料匯出.xlsx",
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/api/admin/resorts/export.xlsx", methods=["GET"])
@require_role("cs")
def admin_export_resorts_xlsx():
    """匯出全部雪場資料為Excel。"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    import tempfile

    conn = get_conn()
    resorts = conn.execute(
        """SELECT r.*, jr.name AS region_name FROM ski_resorts r
           LEFT JOIN japan_regions jr ON r.region_id = jr.id
           ORDER BY jr.display_order, r.name"""
    ).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "雪場資料"
    headers = ["地區", "雪場代碼", "雪場名稱", "狀態", "駐點教練"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E2761")

    for r in resorts:
        coaches = conn.execute(
            """SELECT st.name FROM resort_coaches rc JOIN staff st ON rc.coach_id = st.id
               WHERE rc.resort_id=?""", (r["id"],),
        ).fetchall()
        ws.append([
            r["region_name"], r["code"], r["name"],
            "營運中" if r["is_active"] else "已停用",
            "、".join(c["name"] for c in coaches),
        ])
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)
    conn.close()

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        wb.save(tmp.name)
        path = tmp.name
    return send_file(path, as_attachment=True, download_name="雪場資料匯出.xlsx",
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/api/admin/resorts", methods=["GET"])
@require_role("cs")
def admin_list_resorts():
    conn = get_conn()
    rows = conn.execute("SELECT * FROM ski_resorts ORDER BY id").fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/admin/resorts", methods=["POST"])
@require_role("manager")
def admin_create_resort():
    d = request.json
    conn = get_conn()
    cur = conn.execute(
        "INSERT INTO ski_resorts (region_id, code, name) VALUES (?, ?, ?)",
        (d["region_id"], d["code"], d["name"]),
    )
    resort_id = cur.lastrowid
    conn.execute(
        """INSERT INTO audit_log (staff_id, action, target_type, target_id, before_value, after_value)
           VALUES (?, 'create_resort', 'ski_resort', ?, '{}', ?)""",
        (request.current_staff["id"], resort_id, json.dumps({"name": d["name"], "region_id": d["region_id"]})),
    )
    conn.commit()
    conn.close()
    return jsonify({"id": resort_id}), 201


@app.route("/api/admin/resorts/<int:resort_id>", methods=["PUT"])
@require_role("manager")
def admin_update_resort(resort_id):
    d = request.json
    conn = get_conn()
    before = conn.execute("SELECT * FROM ski_resorts WHERE id=?", (resort_id,)).fetchone()
    if not before:
        conn.close()
        return jsonify({"error": "找不到此雪場"}), 404
    conn.execute(
        "UPDATE ski_resorts SET code=?, name=?, is_active=? WHERE id=?",
        (d.get("code") or before["code"], d.get("name") or before["name"], d.get("is_active", 1), resort_id),
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/resort-coaches", methods=["GET"])
@require_role("cs")
def admin_list_resort_coaches():
    resort_id = request.args.get("resort_id", type=int)
    conn = get_conn()
    q = """SELECT rc.id, rc.resort_id, rc.coach_id, s.name AS coach_name, s.display_code
           FROM resort_coaches rc JOIN staff s ON rc.coach_id = s.id"""
    params = []
    if resort_id:
        q += " WHERE rc.resort_id=?"
        params.append(resort_id)
    rows = conn.execute(q, params).fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/admin/resort-coaches", methods=["POST"])
@require_role("manager")
def admin_assign_resort_coach():
    d = request.json
    conn = get_conn()
    try:
        cur = conn.execute(
            "INSERT INTO resort_coaches (resort_id, coach_id) VALUES (?, ?)",
            (d["resort_id"], d["coach_id"]),
        )
        conn.commit()
        new_id = cur.lastrowid
    except Exception:
        conn.close()
        return jsonify({"error": "此教練已在該雪場名單中"}), 400
    conn.close()
    return jsonify({"id": new_id}), 201


@app.route("/api/admin/resort-coaches/<int:assignment_id>", methods=["DELETE"])
@require_role("manager")
def admin_remove_resort_coach(assignment_id):
    conn = get_conn()
    conn.execute("DELETE FROM resort_coaches WHERE id=?", (assignment_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/coach-schedule", methods=["POST"])
def admin_set_coach_schedule():
    d = request.json
    staff = _current_staff()
    if not staff:
        return jsonify({"error": "未登入或登入已過期,請重新登入"}), 401
    is_self = staff["role"] == "coach" and staff["id"] == d.get("coach_id")
    if not is_self and ROLE_RANK.get(staff["role"], 0) < ROLE_RANK["cs"]:
        return jsonify({"error": "權限不足"}), 403
    conn = get_conn()
    conn.execute(
        """INSERT INTO coach_schedule (coach_id, work_date, status, reason)
           VALUES (?, ?, ?, ?)
           ON CONFLICT(coach_id, work_date) DO UPDATE SET status=excluded.status, reason=excluded.reason""",
        (d["coach_id"], d["work_date"], d["status"], d.get("reason")),
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True}), 201


@app.route("/api/admin/coach/my-bookings", methods=["GET"])
@require_role("coach")
def coach_my_bookings():
    from datetime import date
    today = booking.today_tw().isoformat()
    rows = booking.get_all_bookings(coach_id=request.current_staff["id"], date_from=today)
    return jsonify(rows)


@app.route("/api/admin/coach/my-history", methods=["GET"])
@require_role("coach")
def coach_my_history():
    """教練查詢自己「已上過」的課程歷史(今天以前,不含今天;今天以後的排課在my-bookings)。
    跟my-bookings共用同一支booking.get_all_bookings(),只是把date_from/date_to反過來用。"""
    today = booking.today_tw().isoformat()
    rows = booking.get_all_bookings(coach_id=request.current_staff["id"], date_to=today)
    rows = [r for r in rows if r["date"] != today]
    return jsonify(rows)


@app.route("/api/admin/coach/my-schedule", methods=["GET"])
@require_role("coach")
def coach_my_schedule():
    conn = get_conn()
    rows = conn.execute(
        "SELECT * FROM coach_schedule WHERE coach_id=? ORDER BY work_date",
        (request.current_staff["id"],),
    ).fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/admin/group-classes/check-auto-cancel", methods=["POST"])
@require_role("cs")
def admin_check_group_auto_cancel():
    cancelled = booking.check_and_auto_cancel_group_classes()
    return jsonify({"auto_cancelled_session_ids": cancelled})


@app.route("/api/admin/group-classes", methods=["GET"])
@require_role("cs")
def admin_list_group_classes():
    """列出未來已確認開課(滿2人以上)的團課場次,供後台指派教練用。"""
    from datetime import date
    conn = get_conn()
    rows = conn.execute(
        """SELECT s.*, st.name AS coach_name,
               (SELECT COUNT(*) FROM indoor_session_members sm
                WHERE sm.session_id = s.id AND sm.status='enrolled') AS enrolled_count
           FROM indoor_sessions s LEFT JOIN staff st ON s.coach_id = st.id
           WHERE s.category='group_class' AND s.status IN ('open','confirmed')
           AND s.booking_date >= ?
           ORDER BY s.booking_date, s.start_hour""",
        (booking.today_tw().isoformat(),),
    ).fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/admin/group-classes/<int:session_id>/assign-coach", methods=["POST"])
@require_role("cs")
def admin_assign_group_class_coach(session_id):
    d = request.json
    conn = get_conn()
    before = conn.execute("SELECT coach_id FROM indoor_sessions WHERE id=?", (session_id,)).fetchone()
    if not before:
        conn.close()
        return jsonify({"error": "找不到這個場次"}), 404
    conn.execute("UPDATE indoor_sessions SET coach_id=? WHERE id=?", (d.get("coach_id"), session_id))
    conn.execute(
        """INSERT INTO audit_log (staff_id, action, target_type, target_id, before_value, after_value)
           VALUES (?, 'assign_group_coach', 'indoor_session', ?, ?, ?)""",
        (request.current_staff["id"], session_id,
         json.dumps({"coach_id": before["coach_id"]}), json.dumps({"coach_id": d.get("coach_id")})),
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/coach-schedule", methods=["GET"])
@require_role("cs")
def admin_list_coach_schedule():
    coach_id = request.args.get("coach_id", type=int)
    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")
    conn = get_conn()
    q = "SELECT cs.*, s.name AS coach_name FROM coach_schedule cs JOIN staff s ON cs.coach_id = s.id WHERE 1=1"
    params = []
    if coach_id:
        q += " AND cs.coach_id=?"; params.append(coach_id)
    if date_from:
        q += " AND cs.work_date>=?"; params.append(date_from)
    if date_to:
        q += " AND cs.work_date<=?"; params.append(date_to)
    rows = conn.execute(q, params).fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/admin/dashboard-summary", methods=["GET"])
@require_role("cs")
def admin_dashboard_summary():
    """營運中心首頁彙總資料:今日課程數、報到狀況、當日每位教練應收尾款。
    2026-08:依需求把「待付款訂單」改成「當日每位教練應收尾款」——只抓日本教練課
    (訂金/尾款制)裡,今天出團、尾款還沒收的部分,依教練分組加總,方便教練當天去跟
    學員收現金尾款時對帳;同時把原本「待指派教練」「設備異常」這兩張卡片拿掉
    (共用班表日曆也在同一次改動合併進這個畫面,見前端)。"""
    today_str = booking.today_tw().isoformat()
    conn = get_conn()

    today_sessions = conn.execute(
        "SELECT COUNT(*) c FROM indoor_sessions WHERE booking_date=? AND status != 'cancelled'", (today_str,)
    ).fetchone()["c"]
    today_japan = conn.execute(
        "SELECT COUNT(*) c FROM japan_bookings WHERE booking_date=? AND status != 'cancelled'", (today_str,)
    ).fetchone()["c"]
    today_jump = conn.execute(
        "SELECT COUNT(*) c FROM jump_bookings WHERE booking_date=? AND status != 'cancelled'", (today_str,)
    ).fetchone()["c"]
    total_today = today_sessions + today_japan + today_jump

    checked_in = conn.execute(
        "SELECT COUNT(*) c FROM indoor_sessions WHERE booking_date=? AND attendance_status='completed'", (today_str,)
    ).fetchone()["c"]

    coach_receivables = conn.execute(
        """SELECT jb.coach_id, st.name AS coach_name, SUM(jb.balance_amount) AS amount, COUNT(*) AS booking_count
           FROM japan_bookings jb
           JOIN staff st ON jb.coach_id = st.id
           WHERE jb.booking_date=? AND jb.status != 'cancelled'
             AND jb.balance_paid=0 AND jb.balance_amount > 0
           GROUP BY jb.coach_id, st.name
           ORDER BY amount DESC""",
        (today_str,),
    ).fetchall()

    conn.close()
    return jsonify({
        "today_sessions": total_today,
        "today_checked_in": checked_in,
        "coach_receivables_today": rows_to_dicts(coach_receivables),
    })


@app.route("/api/admin/team-calendar", methods=["GET"])
@require_role("cs")
def admin_team_calendar():
    """
    共用班表日曆:回傳某個月份,每一天所有在職教練的狀態(上班/請假種類)與當天課程數,
    供後台畫出類似TimeTree的「一眼看到誰哪天在忙什麼」共用月曆。
    沒有明確請假紀錄的日期,預設視為正常上班。
    """
    month = request.args.get("month")  # 格式 YYYY-MM
    if not month:
        return jsonify({"error": "請提供 month 參數(YYYY-MM)"}), 400
    year, mo = map(int, month.split("-"))
    import calendar as calendar_module
    last_day = calendar_module.monthrange(year, mo)[1]
    date_from = f"{year:04d}-{mo:02d}-01"
    date_to = f"{year:04d}-{mo:02d}-{last_day:02d}"

    conn = get_conn()
    coaches = conn.execute("SELECT id, name FROM staff WHERE role='coach' AND is_active=1 ORDER BY name").fetchall()

    schedule_rows = conn.execute(
        "SELECT coach_id, work_date, status, reason FROM coach_schedule WHERE work_date >= ? AND work_date <= ?",
        (date_from, date_to),
    ).fetchall()
    schedule_map = {(r["coach_id"], r["work_date"]): {"status": r["status"], "reason": r["reason"]} for r in schedule_rows}

    session_rows = conn.execute(
        """SELECT coach_id, booking_date, COUNT(*) c FROM indoor_sessions
           WHERE status != 'cancelled' AND coach_id IS NOT NULL
             AND booking_date >= ? AND booking_date <= ?
           GROUP BY coach_id, booking_date""",
        (date_from, date_to),
    ).fetchall()
    session_map = {}
    for r in session_rows:
        session_map[(r["coach_id"], r["booking_date"])] = session_map.get((r["coach_id"], r["booking_date"]), 0) + r["c"]

    # 日本教練課:指定教練的,記錄該教練那天去了哪個雪場(一位教練同一天只會有一組,見book_japan_multi_day);
    # 沒有指定教練的(由後台/系統分派),另外依雪場+日期分組,顯示在該天的「未指定教練」清單。
    japan_rows = conn.execute(
        """SELECT jb.coach_id, jb.booking_date, r.name AS resort_name FROM japan_bookings jb
           LEFT JOIN ski_resorts r ON jb.resort_id = r.id
           WHERE jb.status != 'cancelled' AND jb.coach_id IS NOT NULL
             AND jb.booking_date >= ? AND jb.booking_date <= ?""",
        (date_from, date_to),
    ).fetchall()
    japan_resort_map = {}
    for r in japan_rows:
        key = (r["coach_id"], r["booking_date"])
        session_map[key] = session_map.get(key, 0) + 1
        japan_resort_map[key] = r["resort_name"]

    unassigned_japan_rows = conn.execute(
        """SELECT jb.booking_date, r.name AS resort_name, COUNT(*) c FROM japan_bookings jb
           LEFT JOIN ski_resorts r ON jb.resort_id = r.id
           WHERE jb.status != 'cancelled' AND jb.coach_id IS NULL
             AND jb.booking_date >= ? AND jb.booking_date <= ?
           GROUP BY jb.booking_date, r.name""",
        (date_from, date_to),
    ).fetchall()
    unassigned_japan_map = {}
    for r in unassigned_japan_rows:
        unassigned_japan_map.setdefault(r["booking_date"], []).append(
            {"resort_name": r["resort_name"], "count": r["c"]}
        )

    conn.close()

    days = {}
    for day_num in range(1, last_day + 1):
        date_str = f"{year:04d}-{mo:02d}-{day_num:02d}"
        entries = []
        for c in coaches:
            sched = schedule_map.get((c["id"], date_str))
            status = sched["status"] if sched else "working"
            reason = sched["reason"] if sched else None
            entries.append({
                "coach_id": c["id"], "coach_name": c["name"],
                "status": status, "reason": reason,
                "session_count": session_map.get((c["id"], date_str), 0),
                "japan_resort_name": japan_resort_map.get((c["id"], date_str)),
            })
        days[date_str] = {"entries": entries, "unassigned_japan": unassigned_japan_map.get(date_str, [])}

    return jsonify({"month": month, "coaches": rows_to_dicts(coaches), "days": days})


@app.route("/api/booking/japan", methods=["POST"])
def book_japan():
    member_id, err = _require_member_id_from_token()
    if err:
        return err
    d = request.json
    try:
        result = booking.book_japan_multi_day(
            member_id=member_id, bookings=d["bookings"],
            equipment_type=d.get("equipment_type"), participants=d.get("participants"),
            needs_accommodation=d.get("needs_accommodation", False),
            payment_plan=d.get("payment_plan", "full"),
            group_key=d.get("group_key"),
        )
        return jsonify(result), 201
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


# ------------------------------------------------------------------
# 付款
# ------------------------------------------------------------------
_INDOOR_CATEGORY_LABEL = {"trial": "體驗課", "charter": "包機", "self_practice": "自主練習", "group_class": "團課"}


def _log_purchase_notification(conn, member_id, ref_type, ref_id):
    """付款確認成功時呼叫,把「已收到您的訂課/購買」這類訊息記錄成通知,讓會員登入
    會員中心就看得到(2026-08-24新增:之前這幾種購買/訂課完成後完全沒有留下任何
    通知紀錄,會員自己也無從查詢)。查不到對應資料時靜默略過,不影響付款本身。"""
    try:
        if ref_type == "charter_order":
            row = conn.execute(
                "SELECT package_size FROM charter_passes WHERE id=?", (ref_id,)
            ).fetchone()
            if row:
                booking.log_notification(
                    conn, member_id, "purchase_confirmed",
                    f"已收到您購買的包機{row['package_size']}堂堂數包,款項確認完成,堂數已入帳,可以開始訂課。",
                )
        elif ref_type == "indoor_session":
            row = conn.execute(
                "SELECT category, booking_date, start_hour FROM indoor_sessions WHERE id=?", (ref_id,)
            ).fetchone()
            if row:
                label = _INDOOR_CATEGORY_LABEL.get(row["category"], "課程")
                booking.log_notification(
                    conn, member_id, "booking_confirmed",
                    f"已收到您的{label}預約,{row['booking_date']} {row['start_hour']}:00,款項確認完成。",
                )
        elif ref_type == "jump_booking":
            row = conn.execute(
                "SELECT booking_date, start_time FROM jump_bookings WHERE id=?", (ref_id,)
            ).fetchone()
            if row:
                booking.log_notification(
                    conn, member_id, "booking_confirmed",
                    f"已收到您的跳台預約,{row['booking_date']} {row['start_time']},款項確認完成。",
                )
        elif ref_type == "japan_booking":
            row = conn.execute(
                "SELECT resort_id, booking_date FROM japan_bookings WHERE id=?", (ref_id,)
            ).fetchone()
            if row:
                resort = conn.execute(
                    "SELECT name FROM ski_resorts WHERE id=?", (row["resort_id"],)
                ).fetchone()
                resort_name = resort["name"] if resort else ""
                booking.log_notification(
                    conn, member_id, "booking_confirmed",
                    f"已收到您的日本滑雪預約({resort_name} {row['booking_date']}),行程訂金確認完成。",
                )
    except Exception:
        pass  # 通知記錄失敗不該讓付款這個更重要的動作跟著失敗


@app.route("/api/payments/create", methods=["POST"])
def create_payment():
    from payments import active_provider
    member_id, err = _require_member_id_from_token()
    if err:
        return err
    d = request.json
    result = active_provider.create_payment(
        amount=d["amount"], payment_method=d["payment_method"], order_ref=d["order_ref"]
    )
    conn = get_conn()
    try:
        cur = conn.execute(
            """INSERT INTO transactions
               (member_id, ref_type, ref_id, amount, payment_type, payment_method, payment_status, provider_ref)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (member_id, d.get("ref_type"), d.get("ref_id"), d["amount"],
             d.get("payment_type", "full"), d["payment_method"], result["status"], result["provider_ref"]),
        )
        tx_id = cur.lastrowid
        if result["status"] == "confirmed" and d.get("ref_type") == "charter_order":
            booking.finalize_charter_purchase(d["ref_id"], conn=conn)
            _log_purchase_notification(conn, member_id, d.get("ref_type"), d.get("ref_id"))
        elif result["status"] == "confirmed" and d.get("ref_type") in ("indoor_session", "jump_booking", "japan_booking"):
            booking.mark_order_paid(d["ref_type"], d["ref_id"], conn=conn, payment_method=d["payment_method"])
            _log_purchase_notification(conn, member_id, d.get("ref_type"), d.get("ref_id"))
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    return jsonify({"transaction_id": tx_id, **result}), 201


@app.route("/api/admin/sessions/<int:session_id>/cancel", methods=["POST"])
@require_role("cs")
def admin_cancel_session(session_id):
    """客服協調後決定直接取消此場次(取消所有報名此場次的會員,不受時間限制)。"""
    conn = get_conn()
    members = conn.execute(
        "SELECT id FROM indoor_session_members WHERE session_id=? AND status='enrolled'", (session_id,)
    ).fetchall()
    conn.close()
    for m in members:
        try:
            booking.cancel_indoor_booking(m["id"], is_staff=True)
        except ValueError:
            pass
    return jsonify({"ok": True, "cancelled_count": len(members)})


@app.route("/api/admin/check-ins/pending", methods=["GET"])
@require_role("cs")
def admin_pending_check_ins():
    from datetime import date
    up_to_date = request.args.get("up_to_date") or booking.today_tw().isoformat()
    result = booking.get_pending_check_ins(up_to_date)
    return jsonify(result)


@app.route("/api/admin/sessions/<int:session_id>/check-in", methods=["POST"])
@require_role("cs")
def admin_check_in_indoor(session_id):
    d = request.json
    try:
        result = booking.check_in_indoor_session(
            session_id, d["attendance_status"], d.get("lesson_notes"), request.current_staff["id"]
        )
        return jsonify(result)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/admin/jump-bookings/<int:jump_id>/check-in", methods=["POST"])
@require_role("cs")
def admin_check_in_jump(jump_id):
    d = request.json
    try:
        result = booking.check_in_jump_booking(
            jump_id, d["attendance_status"], d.get("lesson_notes"), request.current_staff["id"]
        )
        return jsonify(result)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/admin/japan-bookings/<int:japan_id>/check-in", methods=["POST"])
@require_role("cs")
def admin_check_in_japan(japan_id):
    d = request.json
    try:
        result = booking.check_in_japan_booking(
            japan_id, d["attendance_status"], d.get("lesson_notes"), request.current_staff["id"]
        )
        return jsonify(result)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/admin/sessions/needs-review", methods=["GET"])
@require_role("cs")
def admin_sessions_needs_review():
    """列出因時段衝突而標記「需人工協調」的機台場次。"""
    from datetime import date
    conn = get_conn()
    rows = conn.execute(
        """SELECT s.*, GROUP_CONCAT(m.name, '、') AS member_names
           FROM indoor_sessions s
           JOIN indoor_session_members sm ON sm.session_id = s.id AND sm.status='enrolled'
           JOIN members m ON sm.member_id = m.id
           WHERE s.status='needs_manual_review' AND s.booking_date >= ?
           GROUP BY s.id ORDER BY s.booking_date, s.start_hour""",
        (booking.today_tw().isoformat(),),
    ).fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/admin/sessions/<int:session_id>/resolve", methods=["POST"])
@require_role("cs")
def admin_resolve_session(session_id):
    """客服協調後,標記此場次為已確認保留。"""
    conn = get_conn()
    conn.execute("UPDATE indoor_sessions SET status='confirmed' WHERE id=?", (session_id,))
    conn.execute(
        """INSERT INTO audit_log (staff_id, action, target_type, target_id, before_value, after_value)
           VALUES (?, 'resolve_session_conflict', 'indoor_session', ?, '{"status":"needs_manual_review"}', '{"status":"confirmed"}')""",
        (request.current_staff["id"], session_id),
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/payments/pending", methods=["GET"])
@require_role("cs")
def admin_pending_payments():
    """列出待客服核對入帳的付款(現場付款/匯款轉帳),供後台核准使用。"""
    conn = get_conn()
    rows = conn.execute(
        """SELECT t.*, m.name AS member_name, m.phone AS member_phone
           FROM transactions t JOIN members m ON t.member_id = m.id
           WHERE t.payment_status='awaiting_backoffice_review'
           ORDER BY t.created_at"""
    ).fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/admin/payments/<int:tx_id>/confirm", methods=["POST"])
@require_role("cs")
def confirm_payment(tx_id):
    conn = get_conn()
    try:
        tx = conn.execute("SELECT * FROM transactions WHERE id=?", (tx_id,)).fetchone()
        conn.execute(
            "UPDATE transactions SET payment_status='confirmed', confirmed_by_staff_id=? WHERE id=?",
            (request.current_staff["id"], tx_id),
        )
        conn.execute(
            """INSERT INTO audit_log (staff_id, action, target_type, target_id, before_value, after_value)
               VALUES (?, 'confirm_payment', 'transaction', ?, ?, ?)""",
            (request.current_staff["id"], tx_id,
             json.dumps({"payment_status": tx["payment_status"]}) if tx else "{}",
             json.dumps({"payment_status": "confirmed"})),
        )
        if tx and tx["ref_type"] == "charter_order":
            booking.finalize_charter_purchase(tx["ref_id"], conn=conn)
        elif tx and tx["ref_type"] in ("indoor_session", "jump_booking", "japan_booking"):
            booking.mark_order_paid(tx["ref_type"], tx["ref_id"], conn=conn)
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    return jsonify({"ok": True})


# ------------------------------------------------------------------
# 後台:會員管理 / A-B 方案指派
# ------------------------------------------------------------------
SENSITIVE_MEMBER_FIELDS = ["id_number", "blood_type", "address", "emergency_contact_phone"]


def _mask_sensitive_member_fields(member_dict, staff_role):
    """對照系統分析書13.1「高度敏感資料」分級:客服僅能看到遮罩後的內容,主管以上才看得到完整值。"""
    if ROLE_RANK.get(staff_role, 0) >= ROLE_RANK["manager"]:
        return member_dict
    masked = dict(member_dict)
    for field in SENSITIVE_MEMBER_FIELDS:
        val = masked.get(field)
        if val:
            masked[field] = val[:1] + "*" * max(len(val) - 2, 1) + (val[-1:] if len(val) > 1 else "")
    return masked


@app.route("/api/admin/members", methods=["GET"])
@require_role("cs")
def admin_list_members():
    conn = get_conn()
    rows = conn.execute("SELECT * FROM members ORDER BY created_at DESC").fetchall()
    conn.close()
    members = rows_to_dicts(rows)
    members = [_mask_sensitive_member_fields(m, request.current_staff["role"]) for m in members]
    return jsonify(members)


@app.route("/api/admin/bookings", methods=["GET"])
@require_role("cs")
def admin_list_bookings():
    """訂客資料總覽:彙整所有類型的預約紀錄,可用 query string 篩選
       member_id / category / date_from / date_to"""
    result = booking.get_all_bookings(
        member_id=request.args.get("member_id", type=int),
        category=request.args.get("category"),
        date_from=request.args.get("date_from"),
        date_to=request.args.get("date_to"),
    )
    return jsonify(result)


@app.route("/api/plans/apply", methods=["POST"])
def apply_for_plan():
    member_id, err = _require_member_id_from_token()
    if err:
        return err
    d = request.json
    try:
        result = booking.apply_for_plan(
            member_id=member_id, plan_name=d["plan_name"], billing_cycle=d["billing_cycle"]
        )
        return jsonify(result), 201
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/members/<int:member_id>/plan-applications", methods=["GET"])
@require_member_or_staff()
def member_plan_applications(member_id):
    conn = get_conn()
    rows = conn.execute(
        "SELECT * FROM plan_applications WHERE member_id=? ORDER BY created_at DESC", (member_id,)
    ).fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/admin/plan-applications", methods=["GET"])
@require_role("cs")
def admin_list_plan_applications():
    conn = get_conn()
    rows = conn.execute(
        """SELECT pa.*, m.name AS member_name FROM plan_applications pa
           JOIN members m ON pa.member_id = m.id
           WHERE pa.status='pending' ORDER BY pa.created_at"""
    ).fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/admin/plan-applications/<int:application_id>/review", methods=["POST"])
@require_role("cs")
def admin_review_plan_application(application_id):
    d = request.json
    try:
        result = booking.review_plan_application(
            application_id, request.current_staff["id"], approve=d["approve"], reason=d.get("reason")
        )
        return jsonify(result)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/admin/plan-billing/pending", methods=["GET"])
@require_role("cs")
def admin_pending_plan_billing():
    conn = get_conn()
    rows = conn.execute(
        """SELECT pbr.*, m.name AS member_name FROM plan_billing_records pbr
           JOIN members m ON pbr.member_id = m.id
           WHERE pbr.status='pending' ORDER BY pbr.created_at"""
    ).fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/admin/plan-billing/<int:record_id>/confirm", methods=["POST"])
@require_role("cs")
def admin_confirm_plan_billing(record_id):
    conn = get_conn()
    conn.execute("UPDATE plan_billing_records SET status='paid' WHERE id=?", (record_id,))
    conn.execute(
        """INSERT INTO audit_log (staff_id, action, target_type, target_id, before_value, after_value)
           VALUES (?, 'confirm_plan_billing', 'plan_billing_record', ?, '{"status":"pending"}', '{"status":"paid"}')""",
        (request.current_staff["id"], record_id),
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/members/<int:member_id>/assign-plan", methods=["POST"])
@require_role("cs")
def assign_plan(member_id):
    d = request.json
    conn = get_conn()
    try:
        result = booking.subscribe_plan(
            member_id=member_id, plan_name=d["plan_name"], billing_cycle=d["billing_cycle"],
            assigned_by_staff_id=request.current_staff["id"], conn=conn,
        )
        conn.execute(
            """INSERT INTO audit_log (staff_id, action, target_type, target_id, before_value, after_value)
               VALUES (?, 'assign_plan', 'member', ?, '{}', ?)""",
            (request.current_staff["id"], member_id, json.dumps(result)),
        )
        conn.commit()
        return jsonify(result), 201
    except ValueError as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 400
    finally:
        conn.close()


@app.route("/api/admin/staff", methods=["GET"])
@require_role("manager")
def admin_list_staff():
    conn = get_conn()
    rows = conn.execute(
        "SELECT id, work_id, name, display_code, phone, birthday, role, branch, is_active FROM staff WHERE is_active=1"
    ).fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/admin/staff/<int:staff_id>", methods=["DELETE"])
@require_role("manager")
def admin_delete_staff(staff_id):
    """停用教練/員工帳號(軟刪除:保留歷史預約/稽核紀錄的關聯,僅從清單中隱藏、無法再登入)。"""
    conn = get_conn()
    before = conn.execute("SELECT * FROM staff WHERE id=?", (staff_id,)).fetchone()
    if not before:
        conn.close()
        return jsonify({"error": "找不到此員工"}), 404
    conn.execute("UPDATE staff SET is_active=0 WHERE id=?", (staff_id,))
    conn.execute(
        """INSERT INTO audit_log (staff_id, action, target_type, target_id, before_value, after_value)
           VALUES (?, 'delete_staff', 'staff', ?, ?, ?)""",
        (request.current_staff["id"], staff_id,
         json.dumps({"is_active": 1}), json.dumps({"is_active": 0, "name": before["name"]})),
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/staff", methods=["POST"])
@require_role("manager")
def admin_create_staff():
    """2026-08修正:原本這裡工號重複時,INSERT會因為work_id唯一限制直接丟未攔截的例外,
    回傳一個HTML錯誤頁而不是JSON——前端api()呼叫端解析JSON會失敗、整個新增流程無聲中斷,
    畫面上「新增教練」完全沒有任何成功或失敗的提示,看起來就像「輸入完成但沒有真的新增」。
    改成先檢查工號是否已存在,存在的話回傳清楚的錯誤訊息,不要讓資料庫例外整個往外丟。"""
    d = request.json
    conn = get_conn()
    existing = conn.execute("SELECT id FROM staff WHERE work_id=?", (d["work_id"],)).fetchone()
    if existing:
        conn.close()
        return jsonify({"error": f"工號「{d['work_id']}」已經有人使用,請換一個工號"}), 400
    password = d["birthday"].replace("-", "")[2:8]
    cur = conn.execute(
        """INSERT INTO staff (work_id, name, display_code, phone, birthday, password_hash, role, branch)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
        (d["work_id"], d["name"], d.get("display_code"), d.get("phone"), d["birthday"],
         auth.new_password_hash(password), d["role"], d["branch"]),
    )
    staff_id = cur.lastrowid
    conn.execute(
        """INSERT INTO audit_log (staff_id, action, target_type, target_id, before_value, after_value)
           VALUES (?, 'create_staff', 'staff', ?, '{}', ?)""",
        (request.current_staff["id"], staff_id, json.dumps({"work_id": d["work_id"], "role": d["role"]})),
    )
    conn.commit()
    conn.close()
    return jsonify({"id": staff_id, "default_password": password}), 201


@app.route("/api/auth/demo-staff", methods=["GET"])
def demo_staff_list():
    """僅供本機/測試環境的示範快速登入用(登入頁「員工登入」一鍵按鈕的資料來源)。
    2026-08-24修正:這支API完全沒有驗證身分,卻直接把示範員工的工號跟明碼密碼
    (含老闆帳號)回傳給任何人,正式站(ERSKI_ENV=production)上等於任何訪客都能
    一鍵用老闆身分登入後台——這是一個嚴重的資安漏洞。修法:只有在非正式環境
    才回傳這份清單,正式環境一律回傳空陣列(前端的一鍵登入按鈕列會直接是空的,
    不影響用工號/密碼手動登入)。"""
    if config.IS_PRODUCTION:
        return jsonify([])
    from db import DEMO_STAFF
    return jsonify([
        {"work_id": w, "name": n, "role": r, "password": bday.replace("-", "")[2:8]}
        for w, n, code, bday, r, branch in DEMO_STAFF
    ])


# ------------------------------------------------------------------
# 教練團隊
# ------------------------------------------------------------------
@app.route("/api/coaches", methods=["GET"])
def list_coaches_public():
    """公開:教練團隊介紹頁面(不含證件照)。"""
    conn = get_conn()
    rows = conn.execute(
        """SELECT s.id, s.name, s.display_code, cp.promo_photo, cp.self_intro, cp.resume, cp.experience, cp.rank
           FROM staff s LEFT JOIN coach_profiles cp ON cp.coach_id = s.id
           WHERE s.role='coach' AND s.is_active=1"""
    ).fetchall()
    result = []
    for r in rows:
        certs = conn.execute(
            "SELECT cert_type, cert_name, cert_level FROM coach_certifications WHERE coach_id=?", (r["id"],)
        ).fetchall()
        caps = conn.execute(
            """SELECT co.name FROM coach_capabilities cc
               JOIN coach_capability_options co ON cc.capability_option_id = co.id
               WHERE cc.coach_id=?""",
            (r["id"],),
        ).fetchall()
        d = dict(r)
        d["certifications"] = rows_to_dicts(certs)
        d["capabilities"] = [c["name"] for c in caps]
        result.append(d)
    conn.close()
    return jsonify(result)


@app.route("/api/admin/coaches/<int:coach_id>/basic-info", methods=["GET"])
@require_role("coach")
def admin_get_coach_basic_info(coach_id):
    is_self = request.current_staff["role"] == "coach" and request.current_staff["id"] == coach_id
    if not is_self and ROLE_RANK.get(request.current_staff["role"], 0) < ROLE_RANK["manager"]:
        return jsonify({"error": "權限不足"}), 403
    conn = get_conn()
    row = conn.execute(
        "SELECT id, work_id, name, display_code, phone, birthday, id_number, address, branch, "
        "nickname, email, line_id, instagram, facebook FROM staff WHERE id=?", (coach_id,)
    ).fetchone()
    conn.close()
    if not row:
        return jsonify({"error": "找不到此教練"}), 404
    return jsonify(dict(row))


@app.route("/api/admin/coaches/<int:coach_id>/basic-info", methods=["PUT"])
@require_role("coach")
def admin_update_coach_basic_info(coach_id):
    """教練可以編輯自己的姓名/身分證字號/生日/地址/電話/分店;工號由主管以上異動,避免教練自己誤改登入帳號。"""
    is_self = request.current_staff["role"] == "coach" and request.current_staff["id"] == coach_id
    is_manager_up = ROLE_RANK.get(request.current_staff["role"], 0) >= ROLE_RANK["manager"]
    if not is_self and not is_manager_up:
        return jsonify({"error": "權限不足"}), 403
    d = request.json
    conn = get_conn()
    before = conn.execute("SELECT * FROM staff WHERE id=?", (coach_id,)).fetchone()
    if not before:
        conn.close()
        return jsonify({"error": "找不到此教練"}), 404
    display_code = d.get("display_code") or before["display_code"]
    if not is_manager_up:
        display_code = before["display_code"]  # 顯示代號仍限主管以上異動
    conn.execute(
        "UPDATE staff SET name=?, display_code=?, phone=?, birthday=?, id_number=?, address=?, branch=?, "
        "nickname=?, email=?, line_id=?, instagram=?, facebook=? WHERE id=?",
        (d.get("name") or before["name"], display_code,
         d.get("phone") or before["phone"], d.get("birthday") or before["birthday"],
         d.get("id_number") or before["id_number"], d.get("address") or before["address"],
         d.get("branch") or before["branch"],
         d.get("nickname") if "nickname" in d else before["nickname"],
         d.get("email") if "email" in d else before["email"],
         d.get("line_id") if "line_id" in d else before["line_id"],
         d.get("instagram") if "instagram" in d else before["instagram"],
         d.get("facebook") if "facebook" in d else before["facebook"],
         coach_id),
    )
    conn.execute(
        """INSERT INTO audit_log (staff_id, action, target_type, target_id, before_value, after_value)
           VALUES (?, 'update_coach_basic_info', 'staff', ?, ?, ?)""",
        (request.current_staff["id"], coach_id,
         json.dumps({"name": before["name"], "display_code": before["display_code"],
                     "phone": before["phone"], "branch": before["branch"]}),
         json.dumps(d)),
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/staff/<int:staff_id>/password", methods=["PUT"])
@require_role("coach")
def admin_set_staff_password(staff_id):
    """員工(含教練)變更/重設登入密碼。
    - 本人變更自己的密碼:一定要帶對的current_password才會成功(自助變更密碼,不管是教練
      在/coach專屬頁面、或客服/主管/老闆在員工後台,都是走這支)。
    - 主管以上「代其他員工」重設密碼(例如員工忘記密碼):不需要帶current_password,
      用在後台「教練管理」的「重設密碼」按鈕。
    2026-08新增:之前系統完全沒有任何變更/重設密碼的功能,教練畫面上「若更改生日,登入密碼
    也會跟著變成新生日的六碼」這句提示文字其實是錯的(改生日不會真的改密碼)——這支API連同
    前端「變更密碼」畫面就是補上這個缺口,順便把那句誤導的提示文字改掉。"""
    is_self = request.current_staff["id"] == staff_id
    is_manager_up = ROLE_RANK.get(request.current_staff["role"], 0) >= ROLE_RANK["manager"]
    if not is_self and not is_manager_up:
        return jsonify({"error": "權限不足"}), 403
    d = request.json
    new_password = d.get("new_password")
    require_current = is_self  # 本人一律要驗證目前密碼;主管代他人重設則不用
    try:
        auth.set_staff_password(staff_id, new_password, d.get("current_password"), require_current=require_current)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    conn = get_conn()
    conn.execute(
        """INSERT INTO audit_log (staff_id, action, target_type, target_id, before_value, after_value)
           VALUES (?, 'change_staff_password', 'staff', ?, '{}', '{}')""",
        (request.current_staff["id"], staff_id),
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/location-options", methods=["GET"])
@require_role("coach")
def admin_list_location_options():
    conn = get_conn()
    rows = conn.execute("SELECT * FROM coach_location_options ORDER BY id").fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/admin/location-options", methods=["POST"])
@require_role("manager")
def admin_create_location_option():
    d = request.json
    conn = get_conn()
    try:
        cur = conn.execute(
            "INSERT INTO coach_location_options (name, is_indoor_branch) VALUES (?, ?)",
            (d["name"], 1 if d.get("is_indoor_branch") else 0),
        )
        conn.commit()
        loc_id = cur.lastrowid
    except Exception:
        conn.close()
        return jsonify({"error": "此駐在地名稱已存在"}), 400
    conn.close()
    return jsonify({"id": loc_id}), 201


@app.route("/api/admin/capability-options", methods=["GET"])
@require_role("coach")
def admin_list_capability_options():
    conn = get_conn()
    rows = conn.execute("SELECT * FROM coach_capability_options ORDER BY id").fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/admin/capability-options", methods=["POST"])
@require_role("coach")
def admin_create_capability_option():
    # 2026-08修正:前端教練自助後台的「新增能力選項」按鈕(addCapabilityOptionSelf)本來就是
    # 設計給教練自己用的,但這支API原本被限制成manager以上才能呼叫,教練點下去只會收到
    # 403,按鈕形同壞掉。這是一組公用的能力選項清單(不含機密資料,新增選項不影響任何
    # 已存在的資料),開放給教練也能新增選項跟現有前端設計一致。
    d = request.json
    conn = get_conn()
    try:
        cur = conn.execute("INSERT INTO coach_capability_options (name) VALUES (?)", (d["name"],))
        conn.commit()
        cap_id = cur.lastrowid
    except Exception:
        conn.close()
        return jsonify({"error": "此能力名稱已存在"}), 400
    conn.close()
    return jsonify({"id": cap_id}), 201


@app.route("/api/admin/coaches/<int:coach_id>/details", methods=["GET"])
@require_role("coach")
def admin_get_coach_details(coach_id):
    is_self = request.current_staff["role"] == "coach" and request.current_staff["id"] == coach_id
    if not is_self and ROLE_RANK.get(request.current_staff["role"], 0) < ROLE_RANK["manager"]:
        return jsonify({"error": "權限不足"}), 403
    conn = get_conn()
    profile = conn.execute("SELECT * FROM coach_profiles WHERE coach_id=?", (coach_id,)).fetchone()
    capabilities = conn.execute(
        "SELECT capability_option_id FROM coach_capabilities WHERE coach_id=?", (coach_id,)
    ).fetchall()
    certifications = conn.execute(
        "SELECT id, cert_type, cert_name, cert_level FROM coach_certifications WHERE coach_id=?", (coach_id,)
    ).fetchall()
    locations = conn.execute(
        "SELECT location_option_id FROM coach_locations WHERE coach_id=?", (coach_id,)
    ).fetchall()
    conn.close()
    is_manager_up = ROLE_RANK.get(request.current_staff["role"], 0) >= ROLE_RANK["manager"]
    result = {
        "contract_type": profile["contract_type"] if profile else None,
        "rank": profile["rank"] if profile else None,
        "hourly_rate": profile["hourly_rate"] if profile else None,
        "resume": profile["resume"] if profile else None,
        "experience": profile["experience"] if profile else None,
        "self_intro": profile["self_intro"] if profile else None,
        "years_of_service": profile["years_of_service"] if profile else None,
        "contract_year": profile["contract_year"] if profile else None,
        "discipline": profile["discipline"] if profile else None,
        "specialty": profile["specialty"] if profile else None,
        "snow_years": profile["snow_years"] if profile else None,
        "other_experience": profile["other_experience"] if profile else None,
        "bio_intro": profile["bio_intro"] if profile else None,
        "message_to_students": profile["message_to_students"] if profile else None,
        "coach_motto": profile["coach_motto"] if profile else None,
        "capability_option_ids": [r["capability_option_id"] for r in capabilities],
        "certifications": rows_to_dicts(certifications),
        "location_option_ids": [r["location_option_id"] for r in locations],
    }
    if is_manager_up:
        # 薪資相關資料屬於人事機密,教練自己登入時看不到(即使是查詢自己的details)
        result["base_salary"] = profile["base_salary"] if profile else None
        result["rate_group_class"] = profile["rate_group_class"] if profile else None
        result["rate_trial"] = profile["rate_trial"] if profile else None
        result["rate_assistant"] = profile["rate_assistant"] if profile else None
        result["japan_commission_rate"] = profile["japan_commission_rate"] if profile else None
    return jsonify(result)


@app.route("/api/admin/coaches/<int:coach_id>/details", methods=["PUT"])
@require_role("coach")
def admin_update_coach_details(coach_id):
    """
    一次性覆蓋教練的職稱/能力/證照/駐在地/合約類型/年資/合約年(前端管理整組清單後送出取代)。
    教練自己登入時,能編輯「職稱」「教練能力」「證照」(對外顯示於教練團隊頁面的資料);
    「駐在地(教練分配)」「合約類型」「時薪」「年資」「合約年」屬於人事/派工決策,僅主管以上能異動,
    教練自己送出這幾項會被忽略、維持原值不變。
    """
    is_self = request.current_staff["role"] == "coach" and request.current_staff["id"] == coach_id
    is_manager_up = ROLE_RANK.get(request.current_staff["role"], 0) >= ROLE_RANK["manager"]
    if not is_self and not is_manager_up:
        return jsonify({"error": "權限不足"}), 403
    d = request.json
    conn = get_conn()
    before_profile = conn.execute("SELECT * FROM coach_profiles WHERE coach_id=?", (coach_id,)).fetchone()

    rank = d.get("rank") if "rank" in d else (before_profile["rank"] if before_profile else None)
    if is_manager_up:
        contract_type = d.get("contract_type")
        hourly_rate = d.get("hourly_rate")
        years_of_service = d.get("years_of_service")
        contract_year = d.get("contract_year")
    else:
        contract_type = before_profile["contract_type"] if before_profile else None
        hourly_rate = before_profile["hourly_rate"] if before_profile else None
        years_of_service = before_profile["years_of_service"] if before_profile else None
        contract_year = before_profile["contract_year"] if before_profile else None

    if is_manager_up:
        base_salary = d.get("base_salary")
        rate_group_class = d.get("rate_group_class")
        rate_trial = d.get("rate_trial")
        rate_assistant = d.get("rate_assistant")
        japan_commission_rate = d.get("japan_commission_rate")
    else:
        base_salary = before_profile["base_salary"] if before_profile else None
        rate_group_class = before_profile["rate_group_class"] if before_profile else None
        rate_trial = before_profile["rate_trial"] if before_profile else None
        rate_assistant = before_profile["rate_assistant"] if before_profile else None
        japan_commission_rate = before_profile["japan_commission_rate"] if before_profile else None

    conn.execute(
        f"""INSERT INTO coach_profiles
           (coach_id, contract_type, rank, hourly_rate, years_of_service, contract_year,
            base_salary, rate_group_class, rate_trial, rate_assistant, japan_commission_rate, updated_at)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, {NOW_SQL})
           ON CONFLICT(coach_id) DO UPDATE SET
             contract_type=excluded.contract_type, rank=excluded.rank, hourly_rate=excluded.hourly_rate,
             years_of_service=excluded.years_of_service, contract_year=excluded.contract_year,
             base_salary=excluded.base_salary, rate_group_class=excluded.rate_group_class,
             rate_trial=excluded.rate_trial, rate_assistant=excluded.rate_assistant,
             japan_commission_rate=excluded.japan_commission_rate, updated_at={NOW_SQL}""",
        (coach_id, contract_type, rank, hourly_rate, years_of_service, contract_year,
         base_salary, rate_group_class, rate_trial, rate_assistant, japan_commission_rate),
    )

    if is_manager_up:
        conn.execute("DELETE FROM coach_locations WHERE coach_id=?", (coach_id,))
        for loc_id in d.get("location_option_ids", []):
            conn.execute(
                "INSERT INTO coach_locations (coach_id, location_option_id) VALUES (?, ?) "
                "ON CONFLICT (coach_id, location_option_id) DO NOTHING",
                (coach_id, loc_id),
            )

    conn.execute("DELETE FROM coach_capabilities WHERE coach_id=?", (coach_id,))
    for cap_id in d.get("capability_option_ids", []):
        conn.execute(
            "INSERT INTO coach_capabilities (coach_id, capability_option_id) VALUES (?, ?) "
            "ON CONFLICT (coach_id, capability_option_id) DO NOTHING",
            (coach_id, cap_id),
        )

    conn.execute("DELETE FROM coach_certifications WHERE coach_id=?", (coach_id,))
    for cert in d.get("certifications", []):
        if cert.get("cert_type") and cert.get("cert_level"):
            conn.execute(
                "INSERT INTO coach_certifications (coach_id, cert_type, cert_name, cert_level) VALUES (?, ?, ?, ?)",
                (coach_id, cert["cert_type"], cert.get("cert_name"), cert["cert_level"]),
            )

    conn.execute(
        """INSERT INTO audit_log (staff_id, action, target_type, target_id, before_value, after_value)
           VALUES (?, 'update_coach_details', 'coach', ?, '{}', ?)""",
        (request.current_staff["id"], coach_id, json.dumps(d)),
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/coaches/<int:coach_id>/profile", methods=["GET"])
@require_role("coach")
def admin_get_coach_profile(coach_id):
    is_self = request.current_staff["role"] == "coach" and request.current_staff["id"] == coach_id
    if not is_self and ROLE_RANK.get(request.current_staff["role"], 0) < ROLE_RANK["cs"]:
        return jsonify({"error": "權限不足"}), 403
    conn = get_conn()
    row = conn.execute("SELECT * FROM coach_profiles WHERE coach_id=?", (coach_id,)).fetchone()
    conn.close()
    return jsonify(dict(row) if row else {})


@app.route("/api/admin/coaches/<int:coach_id>/profile", methods=["PUT"])
@require_role("coach")
def admin_update_coach_profile(coach_id):
    is_self = request.current_staff["role"] == "coach" and request.current_staff["id"] == coach_id
    if not is_self and ROLE_RANK.get(request.current_staff["role"], 0) < ROLE_RANK["cs"]:
        return jsonify({"error": "權限不足"}), 403
    d = request.json
    # 自我介紹/給學員一句話/代表教練一句話,前端已限制30字,這裡再做一次後端保險(避免繞過前端直接打API)
    bio_intro = (d.get("bio_intro") or "")[:30] or None if "bio_intro" in d else None
    message_to_students = (d.get("message_to_students") or "")[:30] or None if "message_to_students" in d else None
    coach_motto = (d.get("coach_motto") or "")[:30] or None if "coach_motto" in d else None
    conn = get_conn()
    conn.execute(
        f"""INSERT INTO coach_profiles (coach_id, promo_photo, id_photo, self_intro, resume, experience,
           discipline, specialty, snow_years, other_experience, bio_intro, message_to_students, coach_motto, updated_at)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, {NOW_SQL})
           ON CONFLICT(coach_id) DO UPDATE SET
             promo_photo=COALESCE(excluded.promo_photo, coach_profiles.promo_photo),
             id_photo=COALESCE(excluded.id_photo, coach_profiles.id_photo),
             self_intro=COALESCE(excluded.self_intro, coach_profiles.self_intro),
             resume=COALESCE(excluded.resume, coach_profiles.resume),
             experience=COALESCE(excluded.experience, coach_profiles.experience),
             discipline=COALESCE(excluded.discipline, coach_profiles.discipline),
             specialty=COALESCE(excluded.specialty, coach_profiles.specialty),
             snow_years=COALESCE(excluded.snow_years, coach_profiles.snow_years),
             other_experience=COALESCE(excluded.other_experience, coach_profiles.other_experience),
             bio_intro=COALESCE(excluded.bio_intro, coach_profiles.bio_intro),
             message_to_students=COALESCE(excluded.message_to_students, coach_profiles.message_to_students),
             coach_motto=COALESCE(excluded.coach_motto, coach_profiles.coach_motto),
             updated_at={NOW_SQL}""",
        (coach_id, d.get("promo_photo"), d.get("id_photo"), d.get("self_intro"), d.get("resume"), d.get("experience"),
         d.get("discipline"), d.get("specialty"), d.get("snow_years"), d.get("other_experience"),
         bio_intro, message_to_students, coach_motto),
    )
    conn.execute(
        """INSERT INTO audit_log (staff_id, action, target_type, target_id, before_value, after_value)
           VALUES (?, 'update_coach_profile', 'coach', ?, '{}', ?)""",
        (request.current_staff["id"], coach_id, json.dumps({"self_intro": d.get("self_intro")})),
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


_CERT_FILE_CATEGORIES = ("ski_license", "related_license", "other_license")


@app.route("/api/admin/coaches/<int:coach_id>/certificate-files", methods=["GET"])
@require_role("coach")
def admin_list_coach_certificate_files(coach_id):
    """滑雪證照/相關證照/其他證照的檔案清單(每一類可多筆,圖片或PDF)。權限比照/profile。"""
    is_self = request.current_staff["role"] == "coach" and request.current_staff["id"] == coach_id
    if not is_self and ROLE_RANK.get(request.current_staff["role"], 0) < ROLE_RANK["cs"]:
        return jsonify({"error": "權限不足"}), 403
    conn = get_conn()
    rows = conn.execute(
        "SELECT * FROM coach_certificate_files WHERE coach_id=? ORDER BY category, uploaded_at",
        (coach_id,),
    ).fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/admin/coaches/<int:coach_id>/certificate-files", methods=["POST"])
@require_role("coach")
def admin_upload_coach_certificate_file(coach_id):
    is_self = request.current_staff["role"] == "coach" and request.current_staff["id"] == coach_id
    if not is_self and ROLE_RANK.get(request.current_staff["role"], 0) < ROLE_RANK["cs"]:
        return jsonify({"error": "權限不足"}), 403
    d = request.json
    category = d.get("category")
    file_data = d.get("file_data")
    if category not in _CERT_FILE_CATEGORIES:
        return jsonify({"error": "證照類別不正確"}), 400
    if not file_data:
        return jsonify({"error": "缺少檔案內容"}), 400
    conn = get_conn()
    cur = conn.execute(
        "INSERT INTO coach_certificate_files (coach_id, category, file_name, mime_type, file_data) "
        "VALUES (?, ?, ?, ?, ?)",
        (coach_id, category, d.get("file_name"), d.get("mime_type"), file_data),
    )
    conn.commit()
    new_id = cur.lastrowid
    conn.close()
    return jsonify({"ok": True, "id": new_id})


@app.route("/api/admin/coaches/<int:coach_id>/certificate-files/<int:file_id>", methods=["DELETE"])
@require_role("coach")
def admin_delete_coach_certificate_file(coach_id, file_id):
    is_self = request.current_staff["role"] == "coach" and request.current_staff["id"] == coach_id
    if not is_self and ROLE_RANK.get(request.current_staff["role"], 0) < ROLE_RANK["cs"]:
        return jsonify({"error": "權限不足"}), 403
    conn = get_conn()
    row = conn.execute(
        "SELECT id FROM coach_certificate_files WHERE id=? AND coach_id=?", (file_id, coach_id)
    ).fetchone()
    if not row:
        conn.close()
        return jsonify({"error": "找不到此檔案"}), 404
    conn.execute("DELETE FROM coach_certificate_files WHERE id=?", (file_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


# ------------------------------------------------------------------
# 修改/取消預約(客戶自行操作有時間限制;員工操作不受限)
#
# ⚠️2026-08修正:這幾支路由原本完全沒有member_id參數(只有預約本身的數字id),
# 也沒有檢查操作者跟這筆預約有沒有關係——任何人只要猜到/枚舉一個id就能取消/
# 改期任何人的預約。現在改成:員工(客服以上)不受限;一般使用者則必須帶有效的
# 會員token,而且token裡的member_id要對得上這筆預約實際所屬的會員,否則拒絕。
# ------------------------------------------------------------------
def _is_staff_request():
    staff = _current_staff()
    return bool(staff) and ROLE_RANK.get(staff["role"], 0) >= ROLE_RANK["cs"]


def _check_owns_booking_or_staff(owner_member_id):
    """owner_member_id是這筆預約實際所屬的member_id(None表示預約不存在,留給呼叫端處理404)。
    回傳(is_staff, error_response)——error_response不是None時,呼叫端應直接回傳它。"""
    if _is_staff_request():
        return True, None
    auth_member_id = _current_member_id()
    if auth_member_id is None:
        return False, (jsonify({"error": "未登入或登入已過期,請重新登入"}), 401)
    if owner_member_id is not None and auth_member_id != owner_member_id:
        return False, (jsonify({"error": "無權限操作其他會員的預約"}), 403)
    return False, None


@app.route("/api/booking/indoor/<int:member_ref_id>/cancel", methods=["POST"])
def cancel_indoor(member_ref_id):
    is_staff, err = _check_owns_booking_or_staff(booking.get_indoor_booking_owner(member_ref_id))
    if err:
        return err
    try:
        result = booking.cancel_indoor_booking(member_ref_id, is_staff=is_staff)
        return jsonify(result)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/booking/indoor/<int:member_ref_id>/reschedule", methods=["POST"])
def reschedule_indoor(member_ref_id):
    is_staff, err = _check_owns_booking_or_staff(booking.get_indoor_booking_owner(member_ref_id))
    if err:
        return err
    d = request.json
    try:
        result = booking.reschedule_indoor_booking(
            member_ref_id, d["new_date"], d["new_hour"], is_staff=is_staff
        )
        return jsonify(result)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/booking/jump/<int:jump_id>/cancel", methods=["POST"])
def cancel_jump(jump_id):
    is_staff, err = _check_owns_booking_or_staff(booking.get_jump_booking_owner(jump_id))
    if err:
        return err
    try:
        result = booking.cancel_jump_booking(jump_id, is_staff=is_staff)
        return jsonify(result)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/booking/jump/<int:jump_id>/reschedule", methods=["POST"])
def reschedule_jump(jump_id):
    is_staff, err = _check_owns_booking_or_staff(booking.get_jump_booking_owner(jump_id))
    if err:
        return err
    d = request.json
    try:
        result = booking.reschedule_jump_booking(
            jump_id, d["new_date"], d["new_time"], is_staff=is_staff
        )
        return jsonify(result)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/booking/japan/<group_key>/cancel", methods=["POST"])
def cancel_japan(group_key):
    is_staff, err = _check_owns_booking_or_staff(booking.get_japan_trip_owner(group_key))
    if err:
        return err
    try:
        result = booking.cancel_japan_trip(group_key, is_staff=is_staff)
        return jsonify(result)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/admin/orders/<int:order_id>/discount", methods=["PUT"])
@require_role("manager")
def admin_set_order_discount(order_id):
    d = request.json
    conn = get_conn()
    before = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    if not before:
        conn.close()
        return jsonify({"error": "找不到此訂單"}), 404
    discount_amount = int(d.get("discount_amount", 0))
    if discount_amount < 0 or discount_amount > before["amount"]:
        conn.close()
        return jsonify({"error": "折扣金額不合理"}), 400
    conn.execute("UPDATE orders SET discount_amount=? WHERE id=?", (discount_amount, order_id))
    conn.execute(
        """INSERT INTO audit_log (staff_id, action, target_type, target_id, before_value, after_value)
           VALUES (?, 'set_order_discount', 'order', ?, ?, ?)""",
        (request.current_staff["id"], order_id,
         json.dumps({"discount_amount": before["discount_amount"]}), json.dumps({"discount_amount": discount_amount})),
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/orders/<int:order_id>/record-payment", methods=["POST"])
@require_role("cs")
def admin_record_order_payment(order_id):
    """客服/主管在後台記錄一筆實際收到的款項(可支援訂金/尾款分次收款)。"""
    d = request.json
    conn = get_conn()
    try:
        order = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
        if not order:
            return jsonify({"error": "找不到此訂單"}), 404
        amount = int(d["amount"])
        payment_type = d.get("payment_type", "full")
        payable = order["amount"] - order["discount_amount"]
        new_paid = order["paid_amount"] + amount
        if new_paid > payable:
            return jsonify({"error": f"收款金額超過應付金額(應付NT${payable},已收NT${order['paid_amount']})"}), 400

        conn.execute(
            """INSERT INTO transactions (member_id, order_id, ref_type, ref_id, amount, payment_type, payment_method, payment_status, confirmed_by_staff_id, note)
               VALUES (?, ?, ?, ?, ?, ?, ?, 'confirmed', ?, ?)""",
            (order["member_id"], order_id, order["ref_type"], order["ref_id"], amount,
             payment_type, d.get("payment_method", "onsite"), request.current_staff["id"], d.get("note")),
        )
        new_status = "paid" if new_paid >= payable else order["status"]
        conn.execute("UPDATE orders SET paid_amount=?, status=? WHERE id=?", (new_paid, new_status, order_id))
        conn.execute(
            """INSERT INTO audit_log (staff_id, action, target_type, target_id, before_value, after_value)
               VALUES (?, 'record_order_payment', 'order', ?, ?, ?)""",
            (request.current_staff["id"], order_id,
             json.dumps({"paid_amount": order["paid_amount"]}), json.dumps({"paid_amount": new_paid, "amount": amount})),
        )

        # 若這筆收款讓訂單正式完成付款,依原本邏輯觸發權益產生(與此筆收款合併在同一筆交易內,避免分段寫入不一致)
        if new_status == "paid" and order["status"] != "paid":
            if order["ref_type"] == "charter_pass":
                booking.finalize_charter_purchase(order_id, conn=conn)
            elif order["ref_type"] in ("indoor_session", "jump_booking", "japan_booking"):
                booking.mark_order_paid(order["ref_type"], order["ref_id"], conn=conn)

        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return jsonify({"ok": True, "new_status": new_status, "paid_amount": new_paid})


# 2026-08新增:退款金額達這個門檻(NT$5,000)以上,需要「另一位主管或老闆」二次
# 核准才會真的執行退款,送出申請的人不能自己核准(對照規則書的雙重核准要求)。
REFUND_DUAL_APPROVAL_THRESHOLD = 5000


def _execute_refund(conn, order_id, order, amount, reason, executed_by_staff_id, requested_by_staff_id=None):
    """實際執行退款(寫transactions、更新orders.refunded_amount/status、寫audit_log)。
    金額低於雙重核准門檻時,由originating manager直接呼叫(executed_by_staff_id==本人);
    達門檻時,由核准的第二位主管呼叫,並多帶requested_by_staff_id記錄是誰送的申請。
    呼叫端負責conn.commit()/conn.close()。"""
    conn.execute(
        """INSERT INTO transactions (member_id, order_id, ref_type, ref_id, amount, payment_type, payment_method, payment_status, confirmed_by_staff_id, note)
           VALUES (?, ?, ?, ?, ?, 'refund', 'manual_grant', 'refunded', ?, ?)""",
        (order["member_id"], order_id, order["ref_type"], order["ref_id"], amount,
         executed_by_staff_id, reason),
    )
    new_refunded = order["refunded_amount"] + amount
    new_status = "refunded" if new_refunded >= order["paid_amount"] else order["status"]
    conn.execute(
        """UPDATE orders SET refunded_amount=?, status=?,
           pending_refund_amount=NULL, pending_refund_reason=NULL,
           pending_refund_requested_by=NULL, pending_refund_requested_at=NULL
           WHERE id=?""",
        (new_refunded, new_status, order_id),
    )
    audit_after = {"refunded_amount": new_refunded, "reason": reason}
    if requested_by_staff_id is not None:
        audit_after["requested_by_staff_id"] = requested_by_staff_id
        audit_after["approved_by_staff_id"] = executed_by_staff_id
    conn.execute(
        """INSERT INTO audit_log (staff_id, action, target_type, target_id, before_value, after_value)
           VALUES (?, ?, 'order', ?, ?, ?)""",
        (executed_by_staff_id, "refund_order" if requested_by_staff_id is None else "refund_approved", order_id,
         json.dumps({"refunded_amount": order["refunded_amount"]}),
         json.dumps(audit_after)),
    )
    return new_status, new_refunded


@app.route("/api/admin/orders/<int:order_id>/refund", methods=["POST"])
@require_role("manager")
def admin_refund_order(order_id):
    d = request.json
    conn = get_conn()
    order = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    if not order:
        conn.close()
        return jsonify({"error": "找不到此訂單"}), 404
    amount = int(d["amount"])
    max_refundable = order["paid_amount"] - order["refunded_amount"]
    if amount <= 0 or amount > max_refundable:
        conn.close()
        return jsonify({"error": f"退款金額不合理(最多可退NT${max_refundable})"}), 400
    if order["pending_refund_amount"]:
        conn.close()
        return jsonify({"error": "這筆訂單已經有一筆退款申請待核准,請先處理完再申請新的"}), 400

    if amount < REFUND_DUAL_APPROVAL_THRESHOLD:
        # 小額退款:維持原本行為,manager以上可以直接執行,不用二次核准。
        new_status, new_refunded = _execute_refund(
            conn, order_id, order, amount, d.get("reason"), request.current_staff["id"]
        )
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "status": "refunded", "new_status": new_status, "refunded_amount": new_refunded})

    # 達門檻:這次操作只送出申請,不會真的退款,要等另一位主管/老闆呼叫
    # /refund/approve才會真的執行。
    conn.execute(
        f"""UPDATE orders SET pending_refund_amount=?, pending_refund_reason=?,
           pending_refund_requested_by=?, pending_refund_requested_at={NOW_SQL} WHERE id=?""",
        (amount, d.get("reason"), request.current_staff["id"], order_id),
    )
    conn.execute(
        """INSERT INTO audit_log (staff_id, action, target_type, target_id, before_value, after_value)
           VALUES (?, 'refund_requested', 'order', ?, '{}', ?)""",
        (request.current_staff["id"], order_id,
         json.dumps({"pending_refund_amount": amount, "reason": d.get("reason")})),
    )
    conn.commit()
    conn.close()
    return jsonify({
        "ok": True, "status": "pending_approval",
        "message": f"退款金額達NT${REFUND_DUAL_APPROVAL_THRESHOLD}以上,已送出申請,需由另一位主管或老闆核准後才會實際退款",
    })


@app.route("/api/admin/orders/<int:order_id>/refund/approve", methods=["POST"])
@require_role("manager")
def admin_approve_refund(order_id):
    conn = get_conn()
    order = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    if not order:
        conn.close()
        return jsonify({"error": "找不到此訂單"}), 404
    if not order["pending_refund_amount"]:
        conn.close()
        return jsonify({"error": "這筆訂單目前沒有待核准的退款申請"}), 400
    if order["pending_refund_requested_by"] == request.current_staff["id"]:
        conn.close()
        return jsonify({"error": "不能核准自己送出的退款申請,需要另一位主管或老闆核准"}), 403

    new_status, new_refunded = _execute_refund(
        conn, order_id, order, order["pending_refund_amount"], order["pending_refund_reason"],
        request.current_staff["id"], requested_by_staff_id=order["pending_refund_requested_by"],
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "new_status": new_status, "refunded_amount": new_refunded})


@app.route("/api/admin/orders/<int:order_id>/refund/reject", methods=["POST"])
@require_role("manager")
def admin_reject_refund(order_id):
    d = request.json or {}
    conn = get_conn()
    order = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    if not order:
        conn.close()
        return jsonify({"error": "找不到此訂單"}), 404
    if not order["pending_refund_amount"]:
        conn.close()
        return jsonify({"error": "這筆訂單目前沒有待核准的退款申請"}), 400
    conn.execute(
        """UPDATE orders SET pending_refund_amount=NULL, pending_refund_reason=NULL,
           pending_refund_requested_by=NULL, pending_refund_requested_at=NULL WHERE id=?""",
        (order_id,),
    )
    conn.execute(
        """INSERT INTO audit_log (staff_id, action, target_type, target_id, before_value, after_value)
           VALUES (?, 'refund_rejected', 'order', ?, ?, ?)""",
        (request.current_staff["id"], order_id,
         json.dumps({"pending_refund_amount": order["pending_refund_amount"],
                     "requested_by_staff_id": order["pending_refund_requested_by"]}),
         json.dumps({"rejection_note": d.get("note")})),
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/orders/refunds/pending", methods=["GET"])
@require_role("manager")
def admin_list_pending_refunds():
    conn = get_conn()
    rows = conn.execute(
        """SELECT o.id AS order_id, o.member_id, m.name AS member_name, o.order_type,
                  o.pending_refund_amount, o.pending_refund_reason, o.pending_refund_requested_at,
                  o.pending_refund_requested_by, s.name AS requested_by_name
           FROM orders o
           JOIN members m ON o.member_id = m.id
           LEFT JOIN staff s ON o.pending_refund_requested_by = s.id
           WHERE o.pending_refund_amount IS NOT NULL
           ORDER BY o.pending_refund_requested_at"""
    ).fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


import csv
import io


def _parse_csv(csv_text):
    f = io.StringIO(csv_text)
    return list(csv.DictReader(f))


@app.route("/api/admin/import/members", methods=["POST"])
@require_role("manager")
def admin_import_members():
    """
    匯入既有會員資料。CSV欄位:name,phone,email,birth_date,gender,address,emergency_contact_name,emergency_contact_phone
    僅 name 為必填,其餘欄位可留空。以 phone 或 email 判斷是否已存在(存在則略過,不覆蓋)。
    """
    d = request.json
    try:
        rows = _parse_csv(d["csv_text"])
    except Exception as e:
        return jsonify({"error": f"CSV格式解析失敗:{e}"}), 400

    conn = get_conn()
    created, skipped, failed = [], [], []
    for i, row in enumerate(rows, start=2):  # row 2 = 第一筆資料(第1行是標題)
        name = (row.get("name") or "").strip()
        if not name:
            failed.append({"row": i, "reason": "缺少姓名"})
            continue
        phone = (row.get("phone") or "").strip() or None
        email = (row.get("email") or "").strip() or None
        existing = None
        if phone:
            existing = conn.execute("SELECT id FROM members WHERE phone=?", (phone,)).fetchone()
        if not existing and email:
            existing = conn.execute("SELECT id FROM members WHERE email=?", (email,)).fetchone()
        if existing:
            skipped.append({"row": i, "reason": f"已存在會員(id={existing['id']}),略過"})
            continue
        try:
            cur = conn.execute(
                """INSERT INTO members
                   (name, phone, email, birth_date, gender, address,
                    emergency_contact_name, emergency_contact_phone, auth_provider)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'email')""",
                (name, phone, email, (row.get("birth_date") or "").strip() or None,
                 (row.get("gender") or "").strip() or None, (row.get("address") or "").strip() or None,
                 (row.get("emergency_contact_name") or "").strip() or None,
                 (row.get("emergency_contact_phone") or "").strip() or None),
            )
            created.append({"row": i, "member_id": cur.lastrowid, "name": name})
        except Exception as e:
            failed.append({"row": i, "reason": str(e)})
    conn.commit()
    conn.execute(
        """INSERT INTO audit_log (staff_id, action, target_type, target_id, before_value, after_value)
           VALUES (?, 'import_members', 'members', 0, '{}', ?)""",
        (request.current_staff["id"], json.dumps({"created": len(created), "skipped": len(skipped), "failed": len(failed)})),
    )
    conn.commit()
    conn.close()
    return jsonify({"created": created, "skipped": skipped, "failed": failed})


@app.route("/api/admin/import/coaches", methods=["POST"])
@require_role("manager")
def admin_import_coaches():
    """
    匯入既有教練資料。CSV欄位:name,work_id,birthday,phone,branch
    name/work_id/birthday為必填。預設密碼為生日六碼。以work_id判斷是否已存在(存在則略過)。
    """
    d = request.json
    try:
        rows = _parse_csv(d["csv_text"])
    except Exception as e:
        return jsonify({"error": f"CSV格式解析失敗:{e}"}), 400

    conn = get_conn()
    created, skipped, failed = [], [], []
    for i, row in enumerate(rows, start=2):
        name = (row.get("name") or "").strip()
        work_id = (row.get("work_id") or "").strip()
        birthday = (row.get("birthday") or "").strip()
        if not name or not work_id or not birthday:
            failed.append({"row": i, "reason": "姓名/工號/生日為必填"})
            continue
        existing = conn.execute("SELECT id FROM staff WHERE work_id=?", (work_id,)).fetchone()
        if existing:
            skipped.append({"row": i, "reason": f"工號已存在(id={existing['id']}),略過"})
            continue
        try:
            password = birthday.replace("-", "")[2:8]
            cur = conn.execute(
                """INSERT INTO staff (work_id, name, phone, birthday, password_hash, role, branch)
                   VALUES (?, ?, ?, ?, ?, 'coach', ?)""",
                (work_id, name, (row.get("phone") or "").strip() or None, birthday,
                 auth.new_password_hash(password), (row.get("branch") or "").strip() or "高雄"),
            )
            created.append({"row": i, "staff_id": cur.lastrowid, "name": name, "default_password": password})
        except Exception as e:
            failed.append({"row": i, "reason": str(e)})
    conn.commit()
    conn.execute(
        """INSERT INTO audit_log (staff_id, action, target_type, target_id, before_value, after_value)
           VALUES (?, 'import_coaches', 'staff', 0, '{}', ?)""",
        (request.current_staff["id"], json.dumps({"created": len(created), "skipped": len(skipped), "failed": len(failed)})),
    )
    conn.commit()
    conn.close()
    return jsonify({"created": created, "skipped": skipped, "failed": failed})


@app.route("/api/admin/import/charter-passes", methods=["POST"])
@require_role("manager")
def admin_import_charter_passes():
    """
    匯入既有包機堂數包(讓客戶在舊系統已購買但尚未用完的堂數可以延續)。
    CSV欄位:member_phone_or_email,package_size,headcount_type,remaining
    以電話或Email比對既有會員,找不到會員的那一列會失敗,需先匯入會員。
    """
    d = request.json
    try:
        rows = _parse_csv(d["csv_text"])
    except Exception as e:
        return jsonify({"error": f"CSV格式解析失敗:{e}"}), 400

    conn = get_conn()
    created, failed = [], []
    for i, row in enumerate(rows, start=2):
        key = (row.get("member_phone_or_email") or "").strip()
        if not key:
            failed.append({"row": i, "reason": "缺少會員電話或Email"})
            continue
        member = conn.execute("SELECT id FROM members WHERE phone=? OR email=?", (key, key)).fetchone()
        if not member:
            failed.append({"row": i, "reason": f"找不到對應會員({key}),請先匯入會員資料"})
            continue
        try:
            package_size = int(row["package_size"])
            headcount_type = int(row["headcount_type"])
            remaining = int(row["remaining"])
        except (KeyError, ValueError):
            failed.append({"row": i, "reason": "package_size/headcount_type/remaining 格式錯誤"})
            continue
        cur = conn.execute(
            """INSERT INTO charter_passes (member_id, package_size, headcount_type, remaining)
               VALUES (?, ?, ?, ?)""",
            (member["id"], package_size, headcount_type, remaining),
        )
        conn.execute(
            """INSERT INTO entitlement_ledger
               (member_id, entitlement_type, entitlement_ref_id, change_type, amount, note)
               VALUES (?, 'charter_pass', ?, 'manual_adjust', ?, '資料匯入:延續舊系統堂數')""",
            (member["id"], cur.lastrowid, remaining),
        )
        created.append({"row": i, "member_id": member["id"], "pass_id": cur.lastrowid, "remaining": remaining})
    conn.commit()
    conn.execute(
        """INSERT INTO audit_log (staff_id, action, target_type, target_id, before_value, after_value)
           VALUES (?, 'import_charter_passes', 'charter_passes', 0, '{}', ?)""",
        (request.current_staff["id"], json.dumps({"created": len(created), "failed": len(failed)})),
    )
    conn.commit()
    conn.close()
    return jsonify({"created": created, "failed": failed})


@app.route("/api/admin/orders", methods=["GET"])
@require_role("cs")
def admin_list_orders():
    member_id = request.args.get("member_id", type=int)
    conn = get_conn()
    q = "SELECT o.*, m.name AS member_name FROM orders o JOIN members m ON o.member_id = m.id WHERE 1=1"
    params = []
    if member_id:
        q += " AND o.member_id=?"; params.append(member_id)
    q += " ORDER BY o.created_at DESC"
    rows = conn.execute(q, params).fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/admin/reports/summary", methods=["GET"])
@require_role("manager")
def admin_reports_summary():
    from datetime import date, timedelta
    date_from = request.args.get("date_from") or (booking.today_tw().replace(day=1)).isoformat()
    date_to = request.args.get("date_to") or booking.today_tw().isoformat()
    result = booking.get_report_summary(date_from, date_to)
    return jsonify(result)


@app.route("/api/admin/insurance-brackets", methods=["GET"])
@require_role("manager")
def admin_list_insurance_brackets():
    conn = get_conn()
    rows = conn.execute("SELECT * FROM insurance_brackets ORDER BY bracket_min").fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/admin/insurance-brackets", methods=["PUT"])
@require_role("manager")
def admin_update_insurance_brackets():
    """整批覆蓋勞健保級距表(前端管理整份清單後送出取代),供對照勞保局/健保署最新公告更新用。"""
    d = request.json
    conn = get_conn()
    conn.execute("DELETE FROM insurance_brackets")
    for b in d.get("brackets", []):
        conn.execute(
            """INSERT INTO insurance_brackets
               (bracket_min, bracket_max, insured_salary, labor_insurance_employee, health_insurance_employee)
               VALUES (?, ?, ?, ?, ?)""",
            (b["bracket_min"], b["bracket_max"], b["insured_salary"],
             b["labor_insurance_employee"], b["health_insurance_employee"]),
        )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/coach/my-payroll", methods=["GET"])
@require_role("coach")
def coach_my_payroll():
    """教練查詢自己的薪資紀錄(所有已產生過的月份,最新在前)。薪資紀錄仍然要由主管以上先在
    後台「薪資管理」產生/重新計算過,這裡只負責讀取——不開放教練自己產生,原因是產生流程
    會重算金額並可能覆蓋人工調整過的加班獎金/補貼等欄位,應由主管掌控時機。"""
    conn = get_conn()
    rows = conn.execute(
        "SELECT * FROM coach_payroll_records WHERE coach_id=? ORDER BY period DESC",
        (request.current_staff["id"],),
    ).fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/admin/coach/my-payroll/<int:record_id>/payslip.pdf", methods=["GET"])
@require_role("coach")
def coach_my_payslip(record_id):
    """教練下載自己的薪資單PDF,先確認這筆紀錄真的是自己的才產生,避免猜id看到別人的薪資。"""
    conn = get_conn()
    row = conn.execute(
        "SELECT id FROM coach_payroll_records WHERE id=? AND coach_id=?",
        (record_id, request.current_staff["id"]),
    ).fetchone()
    conn.close()
    if not row:
        return jsonify({"error": "找不到此筆薪資紀錄"}), 404
    import tempfile
    try:
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            output_path = tmp.name
        payroll.generate_payslip_pdf(record_id, output_path)
        conn = get_conn()
        r = conn.execute(
            "SELECT pr.period, st.name FROM coach_payroll_records pr JOIN staff st ON pr.coach_id=st.id WHERE pr.id=?",
            (record_id,),
        ).fetchone()
        conn.close()
        filename = f"薪資單_{r['name']}_{r['period']}.pdf" if r else "薪資單.pdf"
        return send_file(output_path, as_attachment=True, download_name=filename, mimetype="application/pdf")
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/admin/payroll", methods=["GET"])
@require_role("manager")
def admin_list_payroll():
    """查詢某個月份所有教練的薪資紀錄清單(尚未產生過的教練不會出現在清單裡,需先產生)。"""
    period = request.args.get("period")
    if not period:
        return jsonify({"error": "請提供 period 參數(YYYY-MM)"}), 400
    conn = get_conn()
    rows = conn.execute(
        """SELECT pr.*, st.name AS coach_name, cp.rank AS coach_rank FROM coach_payroll_records pr
           JOIN staff st ON pr.coach_id = st.id
           LEFT JOIN coach_profiles cp ON cp.coach_id = st.id
           WHERE pr.period=? ORDER BY st.name""",
        (period,),
    ).fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/admin/payroll/generate", methods=["POST"])
@require_role("manager")
def admin_generate_payroll():
    """為某位教練(或全部教練)產生/重新計算某個月份的薪資紀錄。"""
    d = request.json
    period = d.get("period")
    if not period:
        return jsonify({"error": "請提供 period(YYYY-MM)"}), 400
    coach_id = d.get("coach_id")
    conn = get_conn()
    if coach_id:
        coach_ids = [coach_id]
    else:
        coach_ids = [r["id"] for r in conn.execute("SELECT id FROM staff WHERE role='coach' AND is_active=1").fetchall()]
    conn.close()
    results = [payroll.generate_coach_payroll(cid, period, staff_id=request.current_staff["id"]) for cid in coach_ids]
    return jsonify(results)


@app.route("/api/admin/payroll/<int:record_id>", methods=["PUT"])
@require_role("manager")
def admin_update_payroll_record(record_id):
    """人工更新加班獎金/其他補貼/勞健保覆蓋值/備註,自動重新計算實際所得。"""
    d = request.json
    try:
        result = payroll.update_payroll_manual_fields(
            record_id,
            overtime_bonus=d.get("overtime_bonus"),
            other_subsidy=d.get("other_subsidy"),
            other_subsidy_note=d.get("other_subsidy_note"),
            labor_insurance=d.get("labor_insurance"),
            health_insurance=d.get("health_insurance"),
            notes=d.get("notes"),
            japan_travel_subsidy=d.get("japan_travel_subsidy"),
            japan_transportation_subsidy=d.get("japan_transportation_subsidy"),
        )
        return jsonify(result)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/admin/japan-bookings", methods=["GET"])
@require_role("cs")
def admin_list_japan_bookings():
    """日本教練課訂單總覽,可用 date_from/date_to 篩選,供尾款收款/退佣/轉介管理使用。"""
    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")
    conn = get_conn()
    q = """SELECT jb.*, m.name AS member_name, r.name AS resort_name, st.name AS coach_name,
                  p.name AS rebate_partner_name
           FROM japan_bookings jb
           JOIN members m ON jb.member_id = m.id
           LEFT JOIN ski_resorts r ON jb.resort_id = r.id
           LEFT JOIN staff st ON jb.coach_id = st.id
           LEFT JOIN partner_organizations p ON jb.rebate_partner_id = p.id
           WHERE jb.status != 'cancelled'"""
    params = []
    if date_from:
        q += " AND jb.booking_date >= ?"; params.append(date_from)
    if date_to:
        q += " AND jb.booking_date <= ?"; params.append(date_to)
    q += " ORDER BY jb.booking_date"
    rows = conn.execute(q, params).fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/admin/japan-bookings/<int:booking_id>/collect-balance", methods=["POST"])
@require_role("cs")
def admin_collect_japan_balance(booking_id):
    """後台收取日本教練課尾款(適用於選擇「先繳訂金」的訂單)。"""
    d = request.json
    conn = get_conn()
    jb = conn.execute("SELECT * FROM japan_bookings WHERE id=?", (booking_id,)).fetchone()
    if not jb:
        conn.close()
        return jsonify({"error": "找不到此筆日本教練課訂單"}), 404
    if jb["payment_plan"] != "deposit":
        conn.close()
        return jsonify({"error": "此訂單為全額付款,無需另外收尾款"}), 400
    if not jb["deposit_paid"]:
        conn.close()
        return jsonify({"error": "訂金尚未確認繳納,請先確認訂金"}), 400
    conn.execute(
        """UPDATE japan_bookings SET balance_paid=1, balance_paid_date=?, balance_payment_method=?,
           balance_collected_by_staff_id=? WHERE id=?""",
        (d.get("balance_paid_date") or booking.today_tw().isoformat(), d.get("balance_payment_method"),
         request.current_staff["id"], booking_id),
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/japan-bookings/<int:booking_id>/rebate", methods=["PUT"])
@require_role("manager")
def admin_set_japan_rebate(booking_id):
    """設定退佣(對應合作單位),退介紹費由主管以上手動填寫;會自動重新計算教練收入
    (公式:教練收入 = (報價-退佣金額) x 教練提成比例)。"""
    d = request.json
    conn = get_conn()
    jb = conn.execute("SELECT * FROM japan_bookings WHERE id=?", (booking_id,)).fetchone()
    if not jb:
        conn.close()
        return jsonify({"error": "找不到此筆日本教練課訂單"}), 404
    rebate_partner_id = d.get("rebate_partner_id")
    rebate_amount = d.get("rebate_amount", 0)
    rebate_date = d.get("rebate_date")
    coach_income = jb["coach_income"]
    if jb["coach_commission_rate"] is not None:
        coach_income = round((jb["price"] - rebate_amount) * jb["coach_commission_rate"])
    conn.execute(
        """UPDATE japan_bookings SET rebate_partner_id=?, rebate_amount=?, rebate_date=?, coach_income=?
           WHERE id=?""",
        (rebate_partner_id, rebate_amount, rebate_date, coach_income, booking_id),
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "coach_income": coach_income})


@app.route("/api/admin/japan-bookings/<int:booking_id>/referral", methods=["PUT"])
@require_role("boss")
def admin_set_japan_referral(booking_id):
    """設定轉介單/轉介費用/轉介對象,僅限老闆填寫。"""
    d = request.json
    conn = get_conn()
    jb = conn.execute("SELECT * FROM japan_bookings WHERE id=?", (booking_id,)).fetchone()
    if not jb:
        conn.close()
        return jsonify({"error": "找不到此筆日本教練課訂單"}), 404
    conn.execute(
        "UPDATE japan_bookings SET referral_form=?, referral_fee=?, referral_target=? WHERE id=?",
        (d.get("referral_form"), d.get("referral_fee", 0), d.get("referral_target"), booking_id),
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/partners/<int:partner_id>/japan-bookings", methods=["GET"])
@require_role("manager")
def admin_partner_japan_bookings(partner_id):
    """合作單位底下的日本教練課退佣紀錄:自動帶出訂課人/課程項目/課程人數/課費,退介紹費為手動填寫值。"""
    conn = get_conn()
    rows = conn.execute(
        """SELECT jb.id, jb.booking_date, m.name AS member_name, jb.equipment_type,
                  jb.headcount, jb.price, jb.rebate_amount, jb.rebate_date
           FROM japan_bookings jb JOIN members m ON jb.member_id = m.id
           WHERE jb.rebate_partner_id=? ORDER BY jb.booking_date DESC""",
        (partner_id,),
    ).fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/admin/japan-bookings/<int:booking_id>", methods=["GET"])
@require_role("cs")
def admin_get_japan_booking(booking_id):
    conn = get_conn()
    jb = conn.execute(
        """SELECT jb.*, m.name AS member_name, r.name AS resort_name, st.name AS coach_name,
                  p.name AS rebate_partner_name
           FROM japan_bookings jb
           JOIN members m ON jb.member_id = m.id
           LEFT JOIN ski_resorts r ON jb.resort_id = r.id
           LEFT JOIN staff st ON jb.coach_id = st.id
           LEFT JOIN partner_organizations p ON jb.rebate_partner_id = p.id
           WHERE jb.id=?""",
        (booking_id,),
    ).fetchone()
    conn.close()
    if not jb:
        return jsonify({"error": "找不到此筆日本教練課訂單"}), 404
    return jsonify(dict(jb))


@app.route("/api/admin/profit-loss", methods=["GET"])
@require_role("manager")
def admin_profit_loss():
    period = request.args.get("period")
    if not period:
        return jsonify({"error": "請提供 period 參數(YYYY-MM)"}), 400
    return jsonify(payroll.get_profit_loss_summary(period))


@app.route("/api/admin/payroll/<int:record_id>/payslip.pdf", methods=["GET"])
@require_role("manager")
def admin_download_payslip(record_id):
    import tempfile
    try:
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            output_path = tmp.name
        payroll.generate_payslip_pdf(record_id, output_path)
        conn = get_conn()
        r = conn.execute(
            "SELECT pr.period, st.name FROM coach_payroll_records pr JOIN staff st ON pr.coach_id=st.id WHERE pr.id=?",
            (record_id,),
        ).fetchone()
        conn.close()
        filename = f"薪資單_{r['name']}_{r['period']}.pdf" if r else "薪資單.pdf"
        return send_file(output_path, as_attachment=True, download_name=filename, mimetype="application/pdf")
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/admin/reports/export", methods=["GET"])
@require_role("manager")
def admin_reports_export():
    from datetime import date
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font

    date_from = request.args.get("date_from") or (booking.today_tw().replace(day=1)).isoformat()
    date_to = request.args.get("date_to") or booking.today_tw().isoformat()
    r = booking.get_report_summary(date_from, date_to)

    wb = Workbook()
    bold = Font(bold=True)

    ws1 = wb.active
    ws1.title = "營收總覽"
    ws1.append([f"ERSKI 管理報表 {r['date_from']} ~ {r['date_to']}"])
    ws1["A1"].font = bold
    ws1.append([])
    ws1.append(["項目", "金額/數值"])
    ws1.append(["本期總營收(NT$)", r["total_revenue"]])
    ws1.append(["機台利用率(%)", r["machine_utilization_pct"]])
    ws1.append([])
    ws1.append(["各課程種類營收"])
    ws1.append(["課程種類", "筆數", "金額(NT$)"])
    for x in r["revenue_by_type"]:
        ws1.append([x["label"], x["count"], x["total"]])
    ws1.append([])
    ws1.append(["各課程預約數"])
    ws1.append(["課程種類", "筆數"])
    cat_label = {"trial": "體驗課", "charter": "包機課", "self_practice": "自主練習",
                 "group_class": "團課", "jump": "跳台體驗", "japan": "日本教練課"}
    for k, v in r["booking_counts"].items():
        ws1.append([cat_label.get(k, k), v])

    ws2 = wb.create_sheet("教練績效")
    ws2.append(["教練姓名", "帶課次數", "授課時數", "時薪(NT$)", "預估授課費用(NT$)"])
    for c in r["coach_performance"]:
        ws2.append([c["name"], c["count"], c["hours"], c.get("hourly_rate") or "", c.get("estimated_pay") or ""])

    ws3 = wb.create_sheet("體驗課轉換")
    tc = r["trial_conversion"]
    ws3.append(["體驗課會員數", "已轉正式課程人數", "轉換率(%)"])
    ws3.append([tc["trial_customers"], tc["converted_to_paid"], tc["conversion_rate_pct"]])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=ERSKI_report_{date_from}_{date_to}.xlsx"},
    )


@app.route("/api/admin/entitlement-ledger", methods=["GET"])
@require_role("cs")
def admin_entitlement_ledger():
    member_id = request.args.get("member_id", type=int)
    conn = get_conn()
    q = "SELECT * FROM entitlement_ledger WHERE 1=1"
    params = []
    if member_id:
        q += " AND member_id=?"; params.append(member_id)
    q += " ORDER BY created_at DESC"
    rows = conn.execute(q, params).fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/admin/audit-log", methods=["GET"])
@require_role("manager")
def admin_audit_log():
    conn = get_conn()
    rows = conn.execute("SELECT * FROM audit_log ORDER BY created_at DESC LIMIT 100").fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/faq", methods=["GET"])
def list_faq():
    conn = get_conn()
    rows = conn.execute("SELECT * FROM faq_entries WHERE is_active=1 ORDER BY category, id").fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/faq/ask", methods=["POST"])
def ask_faq():
    """簡易關鍵字比對客服機器人:依關鍵字重疊數量找出最相關的FAQ,找不到則記錄為待回覆問題。"""
    d = request.json
    question_text = (d.get("question") or "").strip()
    if not question_text:
        return jsonify({"error": "請輸入問題內容"}), 400

    conn = get_conn()
    entries = conn.execute("SELECT * FROM faq_entries WHERE is_active=1").fetchall()
    best_match, best_score = None, 0
    for e in entries:
        keywords = [k.strip() for k in (e["keywords"] or "").split(",") if k.strip()]
        score = sum(1 for k in keywords if k in question_text)
        if score > best_score:
            best_score, best_match = score, e

    if best_match:
        conn.close()
        return jsonify({"matched": True, "question": best_match["question"], "answer": best_match["answer"]})

    conn.execute(
        "INSERT INTO faq_unanswered_log (member_id, question_text) VALUES (?, ?)",
        (d.get("member_id"), question_text),
    )
    conn.commit()
    conn.close()
    return jsonify({"matched": False, "message": "很抱歉,目前無法自動回答這個問題,已為你建立客服案件,將由專人協助處理。"})


@app.route("/api/admin/faq", methods=["POST"])
@require_role("cs")
def admin_create_faq():
    d = request.json
    conn = get_conn()
    cur = conn.execute(
        "INSERT INTO faq_entries (question, answer, keywords, category) VALUES (?, ?, ?, ?)",
        (d["question"], d["answer"], d.get("keywords", ""), d.get("category", "")),
    )
    conn.commit()
    faq_id = cur.lastrowid
    conn.close()
    return jsonify({"id": faq_id}), 201


@app.route("/api/admin/faq/<int:faq_id>", methods=["PUT"])
@require_role("cs")
def admin_update_faq(faq_id):
    d = request.json
    conn = get_conn()
    conn.execute(
        "UPDATE faq_entries SET question=?, answer=?, keywords=?, category=?, is_active=? WHERE id=?",
        (d["question"], d["answer"], d.get("keywords", ""), d.get("category", ""), d.get("is_active", 1), faq_id),
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/faq/<int:faq_id>", methods=["DELETE"])
@require_role("cs")
def admin_delete_faq(faq_id):
    conn = get_conn()
    conn.execute("UPDATE faq_entries SET is_active=0 WHERE id=?", (faq_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/faq/unanswered", methods=["GET"])
@require_role("cs")
def admin_list_unanswered_faq():
    conn = get_conn()
    rows = conn.execute(
        """SELECT u.*, m.name AS member_name FROM faq_unanswered_log u
           LEFT JOIN members m ON u.member_id = m.id
           WHERE u.status='pending' ORDER BY u.created_at"""
    ).fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/admin/faq/unanswered/<int:log_id>/resolve", methods=["POST"])
@require_role("cs")
def admin_resolve_unanswered_faq(log_id):
    d = request.json
    conn = get_conn()
    conn.execute(
        "UPDATE faq_unanswered_log SET status='resolved', resolved_by_staff_id=?, resolution_note=? WHERE id=?",
        (request.current_staff["id"], d.get("resolution_note"), log_id),
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/equipment", methods=["GET"])
@require_role("cs")
def admin_list_equipment():
    conn = get_conn()
    rows = conn.execute("SELECT * FROM equipment_items ORDER BY id").fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/admin/equipment", methods=["POST"])
@require_role("manager")
def admin_create_equipment():
    d = request.json
    conn = get_conn()
    cur = conn.execute(
        "INSERT INTO equipment_items (name, equipment_type, notes) VALUES (?, ?, ?)",
        (d["name"], d["equipment_type"], d.get("notes")),
    )
    conn.commit()
    eq_id = cur.lastrowid
    conn.close()
    return jsonify({"id": eq_id}), 201


@app.route("/api/admin/equipment/<int:equipment_id>/status", methods=["PUT"])
@require_role("cs")
def admin_update_equipment_status(equipment_id):
    d = request.json
    conn = get_conn()
    before = conn.execute("SELECT status FROM equipment_items WHERE id=?", (equipment_id,)).fetchone()
    conn.execute("UPDATE equipment_items SET status=? WHERE id=?", (d["status"], equipment_id))
    conn.execute(
        """INSERT INTO audit_log (staff_id, action, target_type, target_id, before_value, after_value)
           VALUES (?, 'update_equipment_status', 'equipment_item', ?, ?, ?)""",
        (request.current_staff["id"], equipment_id,
         json.dumps({"status": before["status"] if before else None}), json.dumps({"status": d["status"]})),
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/equipment/<int:equipment_id>/logs", methods=["GET"])
@require_role("cs")
def admin_list_equipment_logs(equipment_id):
    conn = get_conn()
    rows = conn.execute(
        """SELECT l.*, st.name AS staff_name FROM equipment_maintenance_log l
           LEFT JOIN staff st ON l.staff_id = st.id
           WHERE l.equipment_id=? ORDER BY l.created_at DESC""",
        (equipment_id,),
    ).fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/admin/equipment/<int:equipment_id>/logs", methods=["POST"])
@require_role("cs")
def admin_create_equipment_log(equipment_id):
    d = request.json
    conn = get_conn()
    cur = conn.execute(
        """INSERT INTO equipment_maintenance_log (equipment_id, log_type, description, photo, staff_id)
           VALUES (?, ?, ?, ?, ?)""",
        (equipment_id, d["log_type"], d.get("description"), d.get("photo"), request.current_staff["id"]),
    )
    conn.commit()
    log_id = cur.lastrowid
    conn.close()
    return jsonify({"id": log_id}), 201


@app.route("/api/admin/equipment-logs/<int:log_id>/resolve", methods=["POST"])
@require_role("cs")
def admin_resolve_equipment_log(log_id):
    d = request.json
    conn = get_conn()
    conn.execute(
        f"UPDATE equipment_maintenance_log SET status='resolved', resolved_at={NOW_SQL}, resolution_note=? WHERE id=?",
        (d.get("resolution_note"), log_id),
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/equipment/<int:equipment_id>/closures", methods=["GET"])
@require_role("cs")
def admin_list_equipment_closures(equipment_id):
    conn = get_conn()
    rows = conn.execute(
        "SELECT * FROM equipment_closures WHERE equipment_id=? ORDER BY closure_date", (equipment_id,)
    ).fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/admin/equipment/<int:equipment_id>/closures", methods=["POST"])
@require_role("cs")
def admin_add_equipment_closure(equipment_id):
    d = request.json
    conn = get_conn()
    cur = conn.execute(
        "INSERT INTO equipment_closures (equipment_id, closure_date, reason) VALUES (?, ?, ?)",
        (equipment_id, d["closure_date"], d.get("reason")),
    )
    conn.commit()
    closure_id = cur.lastrowid
    conn.close()
    return jsonify({"id": closure_id}), 201


@app.route("/api/admin/equipment-closures/<int:closure_id>", methods=["DELETE"])
@require_role("cs")
def admin_delete_equipment_closure(closure_id):
    conn = get_conn()
    conn.execute("DELETE FROM equipment_closures WHERE id=?", (closure_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/health", methods=["GET"])
def health():
    return jsonify({"status": "ok"})


if __name__ == "__main__":
    import os
    import db as _db
    if _db.USE_POSTGRES:
        # 正式/測試用PostgreSQL:資料表結構交給Alembic migration管理(見alembic/目錄)，
        # 這裡不再每次啟動都重新執行schema.sql(那樣會在第二次啟動時因為seed資料
        # 已存在而報錯，也不該每次啟動都重新塞測試資料)。
        pass
    elif not os.path.exists(os.path.join(os.path.dirname(__file__), "erski.db")):
        init_db()
    app.run(debug=False, port=5001)
